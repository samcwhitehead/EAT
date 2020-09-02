# -*- coding: utf-8 -*-
"""
GUI to use for post-processing Visual Expresso data

TO DO:
    -Expand functionality to include data from both tracking and channel data 
     (current version just does channel data). To do this, it would be nice to 
     add a function to the main gui that saves a comprehensive summary file
    -Expand list of stats tests and plot types (latter is related to above)
    -Need to (overall) switch everything to csv. this csv and xlsx thing is bad
    -general aesthetic update
    -add in ways to save plots and stats analyses (should the whole package 
    just switch to pandas?)
    -Fix TkDnD for this GUI

Created on Wed Sep 18 10:15:58 2019

@author: Fruit Flies
"""
# -----------------------------------------------------------------------------
import matplotlib

# matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import matplotlib.cm as cm

matplotlib.rcParams['pdf.fonttype'] = 42
matplotlib.rcParams['ps.fonttype'] = 42
import os
import sys
import h5py
import numpy as np
import pandas as pd

import csv
# from statsmodels.stats.multicomp import (MultiComparison, pairwise_tukeyhsd)
import scikit_posthocs as sp
from openpyxl import load_workbook
from scipy.stats import mstats
from sklearn.utils import shuffle

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

if sys.version_info[0] < 3:
    from Tkinter import *
    import tkFileDialog
    from ttk import *
    import tkMessageBox
    import tkSimpleDialog
else:
    from tkinter import *
    from tkinter.ttk import *
    from tkinter import filedialog as tkFileDialog
    from tkinter import messagebox as tkMessageBox
    from tkinter import simpledialog as tkSimpleDialog

from v_expresso_gui_params import (initDirectories, guiParams)
from v_expresso_utils import get_curr_screen_geometry
from bout_and_vid_analysis import (summary_heading, events_heading,
                                   summary_heading_dict, events_heading_dict)

# allows drag and drop functionality. if you don't want this, or are having
#  trouble with the TkDnD installation, set to false.
TKDND_FLAG = True
if TKDND_FLAG:
    from TkinterDnD2 import *

plt.style.use('classic')

# method to use for multiple hypothesis correction
p_adjust_str = 'bonferroni'
# ===================================================================
"""

General utility functions

"""


# -------------------------------------------------------------
# change data format to comply with some posthoc stats tests
def convert_data_for_stats(data_list, data_labels):
    # initialize output array
    data_array = np.array([])
    label_list = []

    # loop through list elements and put data into one array + store labels
    for ith, data_curr in enumerate(data_list):
        data_array = np.append(data_array, data_curr)
        label_list.append(len(data_curr) * [data_labels[ith]])

    label_list_flat = [item for sublist in label_list for item in sublist]
    return data_array, label_list_flat


# --------------------------------------------------------
# align different time series measurment to common t axis
def align_time_series_data(data_list, t_list):
    # initialize output lists
    data_list_out = []
    t_list_out = []

    # loop over datasets (note: 'data_curr' and 't_curr' are sublists)
    for (data_curr, t_curr) in zip(data_list, t_list):
        # get longest time array 
        # N_flies = len(data_curr)
        t_shortest = min(t_curr, key=len)

        # loop through each time series and interp onto longest time vector
        data_interp_list = []
        for (dat, t) in zip(data_curr, t_curr):
            dat_interp = np.interp(t_shortest, t, dat, left=np.nan, right=np.nan)
            data_interp_list.append(dat_interp)

        # concatenate interpolated data arrays together
        data_interp_arr = np.vstack(data_interp_list)

        # add time and data to output list
        data_list_out.append(data_interp_arr)
        t_list_out.append(t_shortest)

    return data_list_out, t_list_out

# -----------------------------------------------------------------------
# helper function to make shuffling data easier (for hypothesis testing)
def my_shuffle(x, N, rand_state=None):
    # perform shuffle
    x_shuffle = shuffle(x, random_state=rand_state)

    # split into arrays of size NxT
    a_out = x_shuffle[:N, :]
    b_out = x_shuffle[N:(2*N), :]

    # return arrays
    return a_out, b_out

# ---------------------------------------------------------------
# hypothesis testing using shuffling (resampling)
def pairwise_shuffle_test(a, b, func=lambda x, y: np.nanmean(np.nanmean(x,axis=0) - np.nanmean(y,axis=0)), N_perm=5000,
                          rand_state=None):
    # subsample data? may speed things up
    a = a[:, ::50]
    b = b[:, ::50]

    # make plot to debug function?
    plotFlag = False

    # we assume that input data arrays a and b are NxT and MxT data arrays,
    # where N,M are the number of entries in a and b, and T is the number of time steps
    N_a = a.shape[0]
    N_b = b.shape[0]

    # get smallest size, since we want to compare consistent group sizes (balanced data)
    N = np.min([N_a, N_b])

    # calculate true difference in data from balanced input arrays
    a_shuf = shuffle(a, random_state=rand_state)[:N, :]
    b_shuf = shuffle(b, random_state=rand_state)[:N, :]
    true_diff = func(a_shuf, b_shuf)

    # combine a and b arrays for shuffling
    X = np.vstack((a, b))

    # perform permutations
    shuffled_diff = np.empty((N_perm, 1))  # initialize storage for shuffled differences
    shuffled_diff[:] = np.nan

    for samp in np.arange(N_perm):
        a_tmp, b_tmp = my_shuffle(X, N)
        shuffled_diff[samp] = func(a_tmp, b_tmp)
        # print(samp)

    # calculate p value using pVal = (k + 1)/(N_perm + 1)
    # where k is number of permutations that yield a more extreme difference than true value
    k = np.sum(np.abs(shuffled_diff) > np.abs(true_diff))
    pVal = (k + 1)/(N_perm + 1)

    if plotFlag:
        fig, ax = plt.subplots(figsize=(8, 4))
        n, bins, patches = ax.hist(shuffled_diff, 200, density=True, histtype='step', cumulative=True,
                                   label='Empirical')
        plt.vlines(true_diff, 0, 1, linestyles="dashed", colors="r")
        fig.show()

    # return pValue
    return pVal

# ---------------------------------------------------------------
# run pairwise permutation tests on multiple (>=2) datasets
def my_shuffle_test(data_list, func=lambda x, y: np.nanmean(np.nanmean(x,axis=0) - np.nanmean(y,axis=0)), N_perm=5000,
                          rand_state=None):
    # get total number of data sets and generate data frame to contain them
    N_dsets = len(data_list)
    pval_mat = np.empty((N_dsets, N_dsets))

    # loop over combinations of datasets
    for ith in range(N_dsets):
        # ones on the diagonal
        pval_mat[ith, ith] = 1.0

        # loop over non-diagonal elements
        for jth in range(ith+1, N_dsets):
            pval_curr = pairwise_shuffle_test(data_list[ith], data_list[jth], func=func, N_perm=N_perm,
                                              rand_state=rand_state)
            pval_mat[ith, jth] = pval_curr
            pval_mat[jth, ith] = pval_curr

    # correct for multiple hypothesis testing (bonferroni)
    N_tests = N_dsets*(N_dsets - 1)/2 
    pval_mat = N_tests*pval_mat
    pval_mat[pval_mat > 1.0] = 1.0

    # convert to pandas dataframe to match output of scikit-posthocs
    pval_df = pd.DataFrame(data=pval_mat)

    # return dataframe
    return pval_df
# ---------------------------------------------------------------
# lighten or darken a color
def lighten_rgb_color(color, amount=0.5):
    """
    Lightens the given color by multiplying (1-luminosity) by the given amount.
    Input can be matplotlib color string, hex string, or RGB tuple.

    Examples:
    >> lighten_color('g', 0.3)
    >> lighten_color('#F034A3', 0.6)
    >> lighten_color((.3,.55,.1), 0.5)
    
    From:
    https://stackoverflow.com/questions/37765197/darken-or-lighten-a-color-in-matplotlib
    """
    import colorsys
    c = colorsys.rgb_to_hls(color[0], color[1], color[2])
    return colorsys.hls_to_rgb(c[0], 1 - amount * (1 - c[1]), c[2])


# =============================================================================
"""

Define frame for loading data. 

This should allow the user to select a summary analysis file (from the main 
visual expresso gui) and load it for plotting/comparison to other groups
 
"""


# =============================================================================

class DataLoader(Frame):
    """ Frame containg buttons to search for data"""

    def __init__(self, parent, col=0, row=0):
        Frame.__init__(self, parent.master)

        # ------------------------------------------------------
        # buttons to allow data loading/clearing
        self.btnframe = Frame(parent.master)
        self.btnframe.grid(column=col, row=row, columnspan=2, sticky=(N, S, E, W))
        # self.btnframe.config(bg='white')

        # section label
        self.frame_label = Label(self.btnframe, text='Select Data:',
                                 font=guiParams['labelfontstr'])
        # background=guiParams['bgcolor'])
        # foreground=guiParams['textcolor'],

        self.frame_label.grid(column=col, row=row, columnspan=3, padx=10, pady=2,
                              sticky=(N, S, E, W))
        # self.btnframe.config(background=guiParams['bgcolor'])

        # -------------------------------------------------------------
        # define load data button
        self.add_button = Button(self.btnframe, text='Load Data',
                                 command=lambda: self.add_data(parent))
        self.add_button.grid(column=col, row=row + 1, padx=10, pady=2,
                             sticky=NSEW)

        # -------------------------------------------------------------
        # define remove current data selection button
        self.rm_button = Button(self.btnframe, text='Remove Data',
                                command=lambda: self.rm_data(parent),
                                state=DISABLED)

        self.rm_button.grid(column=col + 1, row=row + 1, padx=10, pady=2,
                            sticky=NSEW)

        # -------------------------------------------------------------
        # define clear all data files button
        self.clear_button = Button(self.btnframe, text='Clear Data',
                                   command=lambda: self.clear_data(parent))

        self.clear_button.grid(column=col + 2, row=row + 1, padx=10, pady=2,
                               sticky=NSEW)
        # ---------------------------------------------------------------------
        # create listbox to house the different files for potential analysis
        # ---------------------------------------------------------------------        
        self.filelistframe = Frame(parent.master)
        self.filelistframe.grid(column=col, row=row + 1, columnspan=2, padx=10,
                                pady=2, sticky=W)

        self.filelist = Listbox(self.filelistframe, width=64, height=10,
                                selectmode=EXTENDED)

        # --------------------------------------------------------------------                 
        # now make the Listbox and Text drop targets
        if TKDND_FLAG:
            self.filelist.drop_target_register(DND_FILES, DND_TEXT)

        self.filelist.bind('<<ListboxSelect>>', self.on_select)

        if TKDND_FLAG:
            self.filelist.dnd_bind('<<DropEnter>>', postProcess.drop_enter)
            self.filelist.dnd_bind('<<DropPosition>>', postProcess.drop_position)
            self.filelist.dnd_bind('<<DropLeave>>', postProcess.drop_leave)
            self.filelist.dnd_bind('<<Drop>>', postProcess.file_drop)
            self.filelist.dnd_bind('<<Drop:DND_Files>>', postProcess.file_drop)
            self.filelist.dnd_bind('<<Drop:DND_Text>>', postProcess.file_drop)

            self.filelist.drag_source_register(1, DND_TEXT, DND_FILES)
            # text.drag_source_register(3, DND_TEXT)

            # self.filelist.dnd_bind('<<DragInitCmd>>', postProcess.drag_init_listbox)
            # self.filelist.dnd_bind('<<DragEndCmd>>', postProcess.drag_end)
            # text.dnd_bind('<<DragInitCmd>>', drag_init_text)

        # ------------------------------------------------------------
        # scroll bars for listbox
        self.hscroll = Scrollbar(self.filelistframe)
        self.hscroll.pack(side=BOTTOM, fill=X)

        self.vscroll = Scrollbar(self.filelistframe)
        self.vscroll.pack(side=RIGHT, fill=Y)

        self.filelist.config(xscrollcommand=self.hscroll.set,
                             yscrollcommand=self.vscroll.set)
        self.filelist.pack(side=TOP, fill=BOTH)

        self.hscroll.configure(orient=HORIZONTAL,
                               command=self.filelist.xview)
        self.vscroll.configure(orient=VERTICAL,
                               command=self.filelist.yview)

        # index of selected file
        self.selection_ind = []

    # -------------------------------------------------------------------------
    def on_select(self, selection):
        """ Enable or disable based on a valid directory selection. """

        if self.filelist.curselection():
            self.rm_button.configure(state=NORMAL)
        else:
            self.rm_button.configure(state=DISABLED)
            # -------------------------------------------------------------------------

    def add_data(self, parent):
        """ add a new data file into the listbox """
        # get user-selected filename, with user-selected ID name
        new_filename, new_data_name = postProcess.add_data(parent)

        # if we haven't returned a valid filename, just halt here
        if not new_filename or not new_data_name:
            return

        new_entry = "{} -- {}".format(new_data_name, new_filename)
        new_fn_no_ext = os.path.splitext(new_filename)[0]
        # get current listbox entries and their filenames -- want to avoid repetitions of the same file
        entry_list = self.filelist.get(0, END)
        entry_list_fns = [os.path.splitext(ent.split(' -- ')[-1])[0] for ent in entry_list]
        if (len(new_filename) > 0) and (new_entry not in entry_list) and (new_fn_no_ext not in entry_list_fns):
            self.filelist.insert(END, new_entry)
        else:
            print('Not dropping "%s": file does not exist, is invalid, or is a duplicate.' % new_entry)
            return

        # since data list has changed, need to update stats before saving
        parent.statsCalcFlag.set(False)

    # -------------------------------------------------------------------------
    def rm_data(self, parent):
        selected = sorted(self.filelist.curselection(), reverse=True)
        for item in selected:
            self.filelist.delete(item)

        # if we've removed all data, indicate that GUI no longer has data
        if (len(self.filelist.get(0, END)) < 1):
            parent.dataFlag = False

            # if there's a display up, remove the data we just took off
        if parent.displayOnFlag.get():
            postProcess.update_display(parent)

        # since data list has changed, need to update stats before saving
        parent.statsCalcFlag.set(False)

    # -------------------------------------------------------------------------
    def clear_data(self, parent):
        self.filelist.delete(0, END)
        parent.dataFlag = False

        # if there's a display up, remove the data we just took off
        if parent.displayOnFlag.get():
            postProcess.update_display(parent)

        # since data list has changed, need to update stats before saving
        parent.statsCalcFlag.set(False)


# =============================================================================
"""

Define frame for setting data name as well as other plot misc. plot options

 
"""
# =============================================================================
class DataOptions(Frame):
    """ Frame containg buttons to name data sets and set misc. options"""

    def __init__(self, parent, col=0, row=0):
        Frame.__init__(self, parent.master)

        # initialize and place frame
        self.options_frame = Frame(parent.master)
        self.options_frame.grid(column=col, row=row, sticky=(N, S, E, W))

        # initialize boolean check button variable
        self.onlyEatersFlag = BooleanVar()
        self.onlyEatersFlag.set(False)

        # create check button
        self.eater_chkbtn = Checkbutton(self.options_frame,
                                        text="Exclude flies without meals",
                                        variable=self.onlyEatersFlag,
                                        command=lambda: self.update_eater_toggle_var(parent))
        self.eater_chkbtn.grid(column=col, row=row, padx=2, pady=2,
                               sticky=(N, S, E, W))

    # ------------------------------
    # callback functions
    def update_eater_toggle_var(self, parent):
        """ function to update selection of plot variable """
        # read state of checkbutton widget
        chkbtn_val = self.eater_chkbtn.instate(['selected'])
        # read state of checkbutton variable
        curr_val = self.onlyEatersFlag.get()

        # check that widget and variable values agree
        if curr_val != chkbtn_val:
            self.onlyEatersFlag.set(not curr_val)
            curr_val = self.onlyEatersFlag.get()

        # print current state of exclude checkbox variable
        if curr_val:
            print('Only analyzing flies who ate meals')
        else:
            print('Analyzing all flies')

        # reset stats
        parent.statsCalcFlag.set(False)

# =============================================================================
"""

Define frame for setting the variables to plot, as well as the type of 
statistical test to perform on the data 
 
"""


# =============================================================================
class PlotOptions(Frame):
    """ Frame containg buttons to search for data"""

    def __init__(self, parent, col=0, row=0):
        Frame.__init__(self, parent.master)

        # options for plot/stats types
        self.plot_var_choices = parent.plot_var_list
        self.stats_type_choices = parent.stats_type_list

        # ------------------------------------
        # dropdown menu to select plot type
        self.plot_type_frame = Frame(parent.master)
        self.plot_type_frame.grid(column=col, row=row + 1, sticky=(N, S, E, W))

        # label
        self.plot_type_label = Label(self.plot_type_frame,
                                     text='Variable to Plot:',
                                     font=guiParams['labelfontstr'])
        self.plot_type_label.grid(column=col, row=row, padx=2, pady=2,
                                  sticky=(N, S, E, W))

        # initialize menu and populate with choices
        self.plot_option_menu = OptionMenu(self.plot_type_frame,
                                           parent.plotVar,
                                           self.plot_var_choices[0],
                                           *self.plot_var_choices)
        self.plot_option_menu.grid(column=col, row=row + 1, padx=2, pady=2,
                                   sticky=(N, S, E, W))

        # -----------------------------------------
        # dropdown menu to select statistics test
        self.stats_type_frame = Frame(parent.master)
        self.stats_type_frame.grid(column=col, row=row + 2, sticky=(N, S, E, W))
        # label
        self.stats_type_label = Label(self.stats_type_frame,
                                      text='Statistical Test:',
                                      font=guiParams['labelfontstr'])
        self.stats_type_label.grid(column=col, row=row, padx=2, pady=2,
                                   sticky=(N, S, E, W))

        # initialize menu and populate with choices
        self.stats_option_menu = OptionMenu(self.stats_type_frame,
                                            parent.statsType,
                                            self.stats_type_choices[0],
                                            *self.stats_type_choices)
        self.stats_option_menu.grid(column=col, row=row + 1, padx=2, pady=2,
                                    sticky=(N, S, E, W))

        # -----------------------------------------
        # button to update plot window
        self.update_plot_button = Button(parent.master, text='Update Display',
                                         command=lambda: postProcess.update_display(parent))
        self.update_plot_button.grid(column=col + 1, row=row + 1, columnspan=1,
                                     padx=2, pady=2, sticky=(N, S, E, W))

        # -----------------------------------------
        # button to update stats calculation
        self.update_stats_button = Button(parent.master, text='Update Stats',
                                          command=lambda: postProcess.update_stats(parent))
        self.update_stats_button.grid(column=col + 1, row=row + 2, columnspan=1,
                                      padx=2, pady=2, sticky=(N, S, E, W))

        # ---------------------------------------------------------------------
        # callback functions
        def update_plot_var(*args):
            """ function to update selection of plot variable """
            # print(parent.plotVar.get())

        def update_stats_type(*args):
            """ function to update selection of stats type """
            # print(parent.statsType.get())

        # ---------------------------------------------------------
        # bind plot and stats type variables to dropdown menu
        parent.plotVar.trace("w", update_plot_var)
        parent.statsType.trace("w", update_stats_type)


# =============================================================================
"""
Frame for displaying stats, saving, etc
 
"""


# =============================================================================
class SaveOptions(Frame):
    """ Frame containg buttons to save plots and/or stats"""

    def __init__(self, parent, col=0, row=0):
        Frame.__init__(self, parent.master)
        # -----------------------------------------
        # define button to save plot
        self.save_plot_button = Button(parent.master, text='Save Plot',
                                       command=lambda: self.save_plot(parent))
        self.save_plot_button.grid(column=col + 1, row=row, columnspan=1,
                                   padx=2, pady=2, sticky=(N, S, E, W))

        # -----------------------------------------
        # button to update stats calculation
        self.save_stats_button = Button(parent.master, text='Save Stats',
                                        command=lambda: self.save_stats(parent))
        self.save_stats_button.grid(column=col + 2, row=row, columnspan=1,
                                    padx=2, pady=2, sticky=(N, S, E, W))

        # ---------------------------------------------------------------------

    # callback functions
    def save_plot(self, parent):
        """ function to save current plot window """
        if (parent.displayOnFlag.get()):
            fig_save_fn = tkFileDialog.asksaveasfilename(defaultextension=".pdf", title='Select save filename')
            parent.fig.savefig(fig_save_fn)
        else:
            print('No plot to save')

    def save_stats(self, parent):
        """ function to save current stats output """
        if (parent.statsCalcFlag.get()):
            statsType = parent.statsTypeCurr
            stats_save_fn = tkFileDialog.asksaveasfilename(defaultextension=".csv", initialfile=statsType,
                                                           title='Select save filename')
            parent.statsCurr.to_csv(stats_save_fn)

        else:
            print('Need to recalculate stats')


# ==============================================================================
# Main class for GUI
# ==============================================================================
""" 

Main GUI class (called postProcess)

Here we define the layout, size of the full GUI. We also define global 
functions that can be used across different frames. 

"""


class postProcess:

    def __init__(self, master):
        # --------------------------------------------
        # allow references to root (called in main)
        self.master = master

        # --------------------------------------------
        # current data set
        self.data_file_list = []
        self.dataFlag = True

        # --------------------------------------------
        # options for plotting and stats tests
        self.plotVar = StringVar(master)

        # take the plot options from the list of possible summary/event vars
        var_sum = summary_heading
        var_evts = events_heading
        ex_vars = ['Time (s)', 'Filename', 'Bank', 'Channel', 'Meal Number', 'Mealwise Dwell Time Censoring (bool)']
        var_sum = [x for x in var_sum if x not in ex_vars]
        var_evts = [x for x in var_evts if x not in ex_vars]
        plot_var_list = var_sum + var_evts

        self.plot_var_list = plot_var_list

        # hard code in stats options here (for now)
        self.statsType = StringVar(master)
        self.stats_type_list = ['Anderson-Darling', 'Conover', 'Mann-Whitney',
                                'Tukey HSD', 'Wilcoxon', 'Time Series Permutation']

        # just initialize with first entry
        # self.plotVar.set(self.plot_var_list[0])
        # self.statsType.set(self.stats_type_list[0])

        # -------------------------------------------
        # do we currently have a plot displayed?
        self.displayOnFlag = BooleanVar()
        self.displayOnFlag.set(False)

        # do we currently have stats calculated?
        self.statsCalcFlag = BooleanVar()
        self.statsCalcFlag.set(False)

        # --------------------------------------------
        # gui basics
        self.init_gui()

        """ Define blocks within GUI (defined as other classes) """
        # --------------------------------------------
        # initialize instances of frames created above
        self.data_loader = DataLoader(self, col=0, row=0)
        self.data_options = DataOptions(self, col=0, row=2)
        self.plot_options = PlotOptions(self, col=0, row=3)
        self.save_options = SaveOptions(self, col=2, row=4)

        # --------------------------------------------
        # initialize plot window
        # self.canvasFig= plt.figure(1) ;
        self.fig = matplotlib.figure.Figure(figsize=(7, 4), facecolor='none')
        self.ax = self.fig.add_subplot(111)

        self.canvas = matplotlib.backends.backend_tkagg.FigureCanvasTkAgg(self.fig,
                                                                          master=self.master)
        self.canvas.draw()
        self.canvas.get_tk_widget().grid(column=2, row=0, padx=10, pady=5, sticky=N)
        self.canvas._tkcanvas.grid(column=2, row=0, rowspan=4, columnspan=4,
                                   padx=10, pady=5, sticky=N)
        self.fig.tight_layout()
        self.ax.set_axis_off()

        # -------------------------------------------
        # initialize stats storage array
        self.statsCurr = None
        self.statsTypeCurr = None

        # --------------------------------------------
        # extra bit for quit command
        self.master.protocol("WM_DELETE_WINDOW", self.on_quit)

        # ----------------------------------------------
        # make topmost
        self.make_topmost()
    """ functions for main GUI class """

    # ===================================================================
    def on_quit(self):
        """Exits program."""
        if tkMessageBox.askokcancel("Quit", "Do you want to quit?"):
            self.master.destroy()
            self.master.quit()

    # ===================================================================
    def make_topmost(self):
        """Makes this window the topmost window"""
        self.master.lift()
        self.master.attributes("-topmost", 1)
        self.master.attributes("-topmost", 0)
    # ===================================================================
    def init_gui(self):

        # label for GUI        
        self.master.title('Visual Expresso Post Processing')

        # menu bar (to be expanded upon)
        self.master.option_add('*tearOff', 'FALSE')
        self.master.menubar = Menu(self.master)

        # file menu
        self.master.menu_file = Menu(self.master.menubar)
        self.master.menu_file.add_command(label='Exit', command=self.on_quit)

        # add these bits to menu bar
        self.master.menubar.add_cascade(menu=self.master.menu_file, label='File')

    # ===================================================================
    @staticmethod
    def add_data(self):
        """ add summary data file to listbox by requesting user input """
        data_fn = tkFileDialog.askopenfilename(title='Select summary analysis file (file type)')

        # check if data file exists
        if os.path.exists(data_fn):
            self.dataFlag = True

            # also prompt user to assign the data set a name 
            fn_basename = os.path.basename(os.path.splitext(data_fn)[0])
            # print(data_basename)
            user_assigned_name = tkSimpleDialog.askstring("Input", "Assign name for file: {}".format(data_fn),
                                                          initialvalue=fn_basename, parent=self.master)

        # otherwise return empty entries                            
        else:
            data_fn = None
            user_assigned_name = None
        return data_fn, user_assigned_name

    # ===================================================================
    @staticmethod
    def load_data(self):
        """ read in data values from summary file (should separate this into different functions...) """
        data_entry_list = self.data_loader.filelist.get(0, END)
        plot_var = self.plotVar.get()
        onlyEatersFlag = self.data_options.onlyEatersFlag.get()

        # check if we're plotting 'mealwise' or 'summary' data 
        # (this determines how we read xlsx file)
        if "Mealwise" in plot_var:
            mealwiseFlag = True
        else:
            mealwiseFlag = False

        # also check if data is of 'time series' format -- if so, we can't read
        # it from xlsx, and we need make adjustments
        if "Time Series" in plot_var:
            timeSeriesFlag = True
            errStrTS = 'Time series data not contained in {} -- skipping'
        else:
            timeSeriesFlag = False

            # read data filenames/user-assigned data names from entries
        data_fn_list = [s.split(' -- ')[1] for s in data_entry_list]
        data_names_all = [s.split(' -- ')[0] for s in data_entry_list]

        # initialize list to store read data, as well as time array
        data_list = []
        t_list = []
        data_names = []

        # loop over data files
        for ith, fn in enumerate(data_fn_list):

            # get filetype extension and basename
            bn = os.path.basename(fn)
            bn_noExt, ext = os.path.splitext(bn)

            # load data differently depending on filetype
            # ------------------
            # xlsx case
            # ------------------
            if ext == ".xlsx":
                # first check if we're trying to read time series -- this will 
                # not work, so we can skip
                if timeSeriesFlag:
                    print(errStrTS.format(fn))
                    continue

                # load in full xlsx workbook
                wb = load_workbook(fn)

                # restirct attention to either summary or event sheet 
                if mealwiseFlag:
                    try:
                        sheet_index = wb.sheetnames.index('Events')
                        # prefixStr = 'Mealwise '
                    except ValueError:
                        print('Error: no Events data in file')
                        continue
                else:
                    try:
                        sheet_index = wb.sheetnames.index('Summary')
                        # prefixStr = ''
                    except ValueError:
                        print('Error: no Events data in file')
                        continue

                # set selected sheet to active
                wb.active = sheet_index
                sheet = wb.active

                # get column headers to see which data to grab
                sheet_headers = []
                max_col = sheet.max_column
                for c in range(1, max_col + 1):
                    header_str = str(sheet.cell(row=1, column=c).value)
                    # i've changed the way the columns are labeled, but for 
                    # compatibility with previous analysis files, need to 
                    # make sure we include 'mealwise' in name 
                    if mealwiseFlag and not ("Mealwise" in header_str):
                        header_str = "Mealwise " + header_str
                    sheet_headers.append(header_str)

                # try to read data from specified column
                try:
                    col_idx = sheet_headers.index(plot_var)
                except ValueError:
                    print('Could not find variable {}'.format(plot_var))
                    continue

                # how many rows the sheet contains
                max_row = sheet.max_row

                # grab data from that column
                data_curr = np.full((max_row - 1, 1), np.nan)
                for i in range(1, max_row):
                    data_curr[i - 1] = sheet.cell(row=i + 1, column=col_idx + 1).value

                # remove non-eating flies, if selected, as well as nan values
                if onlyEatersFlag and not mealwiseFlag:
                    try:
                        num_meals_col_idx = sheet_headers.index('Number of Meals')
                        num_meals = np.full((max_row - 1, 1), np.nan)
                        for ii in range(1, max_row):
                            num_meals[ii - 1] = sheet.cell(row=ii + 1,
                                                           column=num_meals_col_idx + 1).value

                        ignore_idx = (num_meals < 1) | (np.isnan(data_curr))
                    except ValueError:
                        ignore_idx = np.ones_like(data_curr, dtype=bool)
                else:
                    ignore_idx = np.isnan(data_curr) | (data_curr < 0)

                data_curr = data_curr[~ignore_idx]
                # can't get time series from xlsx, so just make place holder
                t_curr = []
                # ------------------
            # hdf5 case      
            # ------------------
            elif ext == ".hdf5":
                # checking dictionary keys changed between python 2 & 3
                if sys.version_info[0] < 3:
                    summary_key_chk = summary_heading_dict.has_key(plot_var)
                    events_key_chk = events_heading_dict.has_key(plot_var)
                else:
                    summary_key_chk = plot_var in summary_heading_dict
                    events_key_chk = plot_var in events_heading_dict

                # get dataset name for current plot variable
                if summary_key_chk:
                    dset_name = summary_heading_dict[plot_var]
                    # because h5py no longer supports .value for grabbing data,
                    # need to read out data differently if it's scalar or array
                    if timeSeriesFlag:
                        scalarFlag = False
                    else:
                        scalarFlag = True
                elif events_key_chk:
                    dset_name = events_heading_dict[plot_var]
                    scalarFlag = False
                else:
                    errStrH5 = 'Could not find {} data in {} -- skipping'
                    print(errStrH5.format(plot_var, fn))
                    continue

                # read out all group keys in hdf5 file (these correspond to individual flies)
                with h5py.File(fn, 'r') as f:
                    fly_grp_list = f.keys()
                    data_curr = []
                    t_curr = []

                    # --------------------------------------------------------------
                    # perform checks so that non-feeding data doesn't hit an error
                    # check if this data file contains eating data. if not, skip it
                    dset_chk_list = [("{}/{}".format(grp, dset_name) in f) for grp in fly_grp_list]
                    num_meals_chk_list = [("{}/{}".format(grp, 'num_meals') in f) for grp in fly_grp_list]
                    if not any(dset_chk_list):
                        print('{} does not contain {} data -- skipping'.format(fn, plot_var))
                        continue
                    elif not any(num_meals_chk_list) and onlyEatersFlag:
                        chkbtn_str = "Exclude flies without meals"
                        err_str = '{} does not contain meal data, and <<{}>> has been selected -- skipping'
                        print(err_str.format(os.path.normpath(fn), chkbtn_str))
                        continue
                    # ---------------------------------------------------------
                    # loop over groups (flies) and read out data
                    for grp in fly_grp_list:
                        # skip if fly doesn't eat and we're supposed to skip non-eating flies
                        if onlyEatersFlag:
                            num_meals = f[grp]['num_meals'][()]

                            # if not meals present, skip current file
                            if num_meals < 1:
                                continue

                        # read out data for current plot variable (need to read out differently depending on format)
                        arrayCheck = isinstance(f[grp][dset_name], (list, tuple, np.ndarray))
                        if scalarFlag or not arrayCheck:
                            h5_dat = f[grp][dset_name][()]
                        else:
                            h5_dat = f[grp][dset_name][:]

                        data_curr.append(h5_dat)

                        # if time series case, also grab time
                        if timeSeriesFlag:
                            t_dat = f[grp]['t'][:]
                            t_curr.append(t_dat)
                        elif mealwiseFlag and not timeSeriesFlag:
                            # if we're not using time series data, but don't 
                            # have scalar data, need to flatten list
                            if len(data_curr) == 1:
                                # if there's only one entry in "data_curr", just convert to list
                                data_curr = list(data_curr)
                            elif isinstance(data_curr[0], (np.float64, np.ndarray)):
                                # if data is numpy type, squeeze entries in the "data_curr" list to one dim
                                data_curr_squeeze = [np.squeeze(d) for d in data_curr]

                                # then concatenate entries into one array and convert to list
                                data_curr = list(np.hstack(data_curr_squeeze))
                            else:
                                data_curr = [item for sublist in data_curr for item in sublist]

                    # remove nan values from data
                    if timeSeriesFlag:
                        data_curr = [d for d in data_curr if not np.all(np.isnan(d))]
                    else:
                        data_curr = [d for d in data_curr if not np.isnan(d)]
            # ------------------
            # csv case      
            # ------------------
            elif ext == ".csv":
                print("under construction")
                continue
            # ------------------
            # unknown case
            # ------------------
            else:
                print("Invalid data file")
                continue

            # --------------------------------------------
            # add data from current file to list of data
            data_list.append(data_curr)
            t_list.append(t_curr)
            data_names.append(data_names_all[ith])

        # -----------------------------------------------
        # check for empty data sets, and remove those
        # empty_idx = [(not d) for d in data_list]
        # data_list = data_list[not empty_idx]
        # data_names = data_names[not empty_idx]
        # t_list = t_list[not empty_idx]

        return data_list, data_names, t_list

    # ===================================================================
    @staticmethod
    def update_display(self):
        """ update inline plot display """

        # get currently loaded data
        (data_list, data_names, t_list) = postProcess.load_data(self)
        N_dsets = len(data_list)
        if (N_dsets < 1):
            print('No data to plot')
            self.displayOnFlag.set(False)
            self.ax.cla()
            self.canvas.draw()
            return

        # set of colors used for bar plots
        # cmap = cm.get_cmap('Pastel1')
        # norm_range = np.linspace(0, 1, num=N_dsets)
        # colors = cmap(norm_range)
        colors = plt.cm.Pastel1(range(50))

        # get current plotting data
        plot_var = self.plotVar.get()

        # see if we're working with time series data
        if "Time Series" in plot_var:
            timeSeriesFlag = True
        else:
            timeSeriesFlag = False

            # clear previous plot
        self.ax.cla()

        # -----------------------------------------
        # switch plot types depending on data type
        # -----------------------------------------
        if timeSeriesFlag:
            # assume for now that we want to calculate the mean and SEM for
            # the time series data. to do this, we need to first align the
            # data to a common time axis
            data_align, t_align = align_time_series_data(data_list, t_list)
            # initialize lists so we can keep track of x axis limits
            t_max_list = []
            t_min_list = []

            # loop over data sets and plot
            for ith, (data, t) in enumerate(zip(data_align, t_align)):
                # get mean, STD, and SEM of current data
                N_flies = data.shape[0]
                data_mean = np.nanmean(data, axis=0)
                data_std = np.nanstd(data, axis=0)
                data_se = data_std / np.sqrt(N_flies)

                # subsample data. NB: this is only for plotting, but makes drawing to canvas much faster
                sub_samp_step = 50
                t = t[::sub_samp_step]
                data_mean = data_mean[::sub_samp_step]
                data_se = data_se[::sub_samp_step]

                # current plot color
                color_curr = lighten_rgb_color(colors[ith], 1.6)

                # plot confidence interval (mean +/- SEM)
                self.ax.fill_between(t, data_mean + data_se,
                                     data_mean - data_se,
                                     color=color_curr,
                                     alpha=0.5)

                # plot mean
                self.ax.plot(t, data_mean, lw=2, label=data_names[ith],
                             color=color_curr)

                # get current min/max time values
                t_max_list.append(np.max(t))
                t_min_list.append(np.min(t))

            # -------------------------------
            # axis properties
            self.ax.legend(loc='upper left', fontsize='small')
            self.ax.set_xlabel('Time (s)')  # label x axis
            self.ax.set_xlim(left=min(t_min_list), right=max(t_max_list))

        else:
            # box plot properties
            boxprops = dict(linestyle='-', linewidth=1, color='k')
            flierprops = dict(marker='o', markeredgecolor='k', markersize=6,
                              markerfacecolor='none', linestyle='none')
            medianprops = dict(linestyle='-', linewidth=1.5, color='k')
            whiskerprops = dict(linestyle='-', linewidth=1, color='k')
            # plot new data
            self.bplot = self.ax.boxplot(data_list,
                                         vert=True,  # vertical box alignment
                                         patch_artist=True,  # fill with color
                                         labels=data_names,  # to label boxes
                                         boxprops=boxprops,  # additional box properties
                                         flierprops=flierprops,
                                         medianprops=medianprops,
                                         whiskerprops=whiskerprops)

            # ---------------------------
            # plot and axis props
            for patch, color in zip(self.bplot['boxes'], colors):
                patch.set_facecolor(color)
            # handle axis labels, title, etc
            self.ax.set_xticklabels(data_names, fontsize='small', rotation=25)

        # ---------------------------------------------------
        # plot options not unique to plot type
        self.ax.set_ylabel(plot_var)  # y axis label
        self.ax.autoscale(enable=True, axis='y', tight=False)  # correct y lim
        self.fig.tight_layout()  # make sure axis labels fit

        # draw figure
        self.canvas.draw()

        # update the boolean value that tells us whether the display is up
        self.displayOnFlag.set(True)

        # update stats boolean to make sure we're not using the wrong stats
        self.statsCalcFlag.set(False)

    # ===================================================================
    @staticmethod
    def update_stats(self):
        """ update statstics test """
        # get currently loaded data
        (data_list, data_names, t_list) = postProcess.load_data(self)

        # make sure we're comparing at least two groups
        if len(data_list) < 2:
            print('Need to compare at least two data sets')
            return

        # get name of current data (to check if we're working with time series values or not)
        plot_var = self.plotVar.get()
        timeSeriesFlag = ("Time Series" in plot_var)

        # get type of stats test
        stats_type = self.statsType.get()

        # if we're using time series data, make sure selected test is 'Time Series Shuffle'
        incompatibleFlag = (timeSeriesFlag and not (stats_type == 'Time Series Permutation')) or \
                           (not timeSeriesFlag and (stats_type == 'Time Series Permutation'))
        if incompatibleFlag:
            print('Cannot use {} test with {} data'.format(stats_type, plot_var))
            return

        # ------------------------------------------------------------------------
        # run stats test (plan is to fill this in with more test as time goes on)
        if (stats_type == 'Anderson-Darling'):
            pval_mat = sp.posthoc_anderson(data_list, p_adjust=p_adjust_str)
        elif (stats_type == 'Conover'):
            pval_mat = sp.posthoc_conover(data_list, p_adjust=p_adjust_str)
        elif (stats_type == 'Mann-Whitney'):
            pval_mat = sp.posthoc_mannwhitney(data_list, p_adjust=p_adjust_str)
        # elif (stats_type == 'Tukey'):
        #     # convert data to array
        #    (data, labels) = convert_data_for_stats(data_list, data_names)
        #    pval_mat = sp.posthoc_tukey(data)
        elif (stats_type == 'Tukey HSD'):
            # convert data to array
            (data, labels) = convert_data_for_stats(data_list, data_names)
            pval_mat = sp.posthoc_tukey_hsd(data, labels, p_adjust=p_adjust_str)
            print('NB: for Tukey HSD, 1 = significant, 0 = not significant,' +
                  'and -1 = diagonal element at alpha = 0.05')
        elif (stats_type == 'Wilcoxon'):
            try:
                pval_mat = sp.posthoc_wilcoxon(data_list, p_adjust=p_adjust_str)
            except ValueError:
                print('Unequal N -- cannot use Wilcoxon test')
                return
        elif (stats_type == 'Time Series Permutation'):
            # since using time series data, need to align data sets
            data_align, t_align = align_time_series_data(data_list, t_list)

            # get shortest time list -- restrict all data to this range
            t_len_list = [t.size for t in t_align]
            t_min_ind = min(t_len_list)

            data_align_trim = [data[:, :t_min_ind] for data in data_align]

            # apply hypothesis test to data pairs
            pval_mat = my_shuffle_test(data_align_trim)
        else:
            print('Invalid hypothesis test selection')
            return

        # assign labels to data frame output
        pval_mat.columns = data_names
        pval_mat.index = data_names

        # print stats results (need to move this to gui)
        print(pval_mat)

        # update title to display stats
        #         self.ax.set_title('{} test comparing {}: p = {}'.format(stats_type,
        #                           plot_var, pval), fontsize='small')
        # make sure axis labels fit
        self.fig.tight_layout()
        # draw
        self.canvas.draw()

        # add current stats to GUI structure
        self.statsCalcFlag.set(True)
        self.statsCurr = pval_mat
        self.statsTypeCurr = stats_type

    # =====================================================================
    """ TkDND functions """
    if TKDND_FLAG:
        @staticmethod
        def drop_enter(event):
            event.widget.focus_force()
            # print('Entering widget: %s' % event.widget)
            # print_event_info(event)
            return event.action

        @staticmethod
        def drop_position(event):
            # print('Position: x %d, y %d' %(event.x_root, event.y_root))
            # print_event_info(event)
            return event.action

        @staticmethod
        def drop_leave(event):
            # print('Leaving %s' % event.widget)
            # print_event_info(event)
            return event.action

        # specific functions for different listboxes
        @staticmethod
        def file_drop(event):
            if event.data:
                valid_ext = [".xlsx", ".hdf5"]  # [".xlsx", ".csv"]

                files = event.widget.tk.splitlist(event.data)
                for f in files:
                    if os.path.exists(f) and f.endswith(tuple(valid_ext)):

                        # also prompt user to assign the data set a name 
                        fn_basename = os.path.basename(os.path.splitext(f)[0])
                        data_name = tkSimpleDialog.askstring("Input",
                                                             "Assign name for file: {}".format(f),
                                                             initialvalue=fn_basename)
                        new_entry = "{} -- {}".format(data_name, f)

                        # check to see if new item is already in listbox. if so, don't add
                        fn_no_ext = os.path.splitext(f)[0]  # filename without extension
                        entry_list = event.widget.get(0, END)  # current list of entries
                        # filenames (without extensions) for current listbox entries
                        entry_list_fns = [os.path.splitext(ent.split(' -- ')[-1])[0] for ent in entry_list]

                        if (new_entry not in entry_list) and (fn_no_ext not in entry_list_fns):
                            event.widget.insert('end', new_entry)
                        else:
                            print('Not dropping file "%s": file seems to be a duplicate of extant entry.' % f)

                    else:
                        print('Not dropping file "%s": file does not exist or is invalid.' % f)

            return event.action


# ==============================================================================
""" Run main loop """
if __name__ == '__main__':
    screen_geometry = get_curr_screen_geometry()
    # print(screen_geometry)
    if TKDND_FLAG:
        root = TkinterDnD.Tk()
    else:
        root = Tk()
    postProcess(root)
    root.mainloop()
