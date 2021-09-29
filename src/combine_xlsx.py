# -*- coding: utf-8 -*-
"""
Quick script for combining data in multiple xlsx files to produce a

Created on Friday Sept 17 11:55 2021

@author: Fruit Flies
"""
# -----------------------------------------------------------
# import relevant modules
import os
import re
import numpy as np
from openpyxl import Workbook, load_workbook

# params
# *** update "data_path" depending on where xlsx files are stored
data_path = "C:/Users/samcw/Dropbox/Paper Manuscripts/Expresso paper/figure_data/Figure7/xls/"
save_fn_name = "combined"  # prefix for combined filename
fn_search_exp = "(?P<starvePeriod>\d+)hr_(?P<foodType>\d+)(?P<foodTypeUnits>[mM]+)"  # info on filenames for xlsx files (what to look for)
pat = re.compile(fn_search_exp)

# -----------------------------------------------------------
# function to load info from a given xlsx file
def my_load_xlsx(filename_full):
    # load current workbook
    wb = load_workbook(filename_full)

    # initialize lists for output
    data_headers_list = []
    data_out_list = []

    # get sheet names in workbook
    sheet_names = wb.sheetnames

    # loop over sheet names and read out data
    for sht in sheet_names:
        # make the current sheet active
        sheet_index = wb.sheetnames.index(sht)
        wb.active = sheet_index
        sheet = wb.active

        # how many rows the sheet contains
        max_row = sheet.max_row

        # read out column names from sheet
        col_names = []
        for cell in sheet[1]:
            col_names.append(cell.value)

        # append column names (with sheet info) to output list
        headers_curr = [sht + "--" + col for col in col_names]
        data_headers_list.extend(headers_curr)


        # read out data from each column
        for col in col_names:
            # get index for current data column
            col_idx = col_names.index(col)

            # read data from column
            data_curr = []
            for i in range(1, max_row):
                data_curr.append(sheet.cell(row=i + 1, column=col_idx + 1).value)

            # when reading from cells, each cell gives an array. want just entries
            # data_curr = np.asarray([d[0] for d in data_curr])

            # append current data to output list
            data_out_list.append(data_curr)

    # return data and header info
    return data_headers_list, data_out_list


######################################################################
""" MAIN SCRIPT """
if __name__ == '__main__':
    # path info
    data_path = os.path.normpath(data_path)  # try to make it easier for switching between windows and Unix
    file_ext = ".xlsx"
    save_fn = save_fn_name + file_ext

    # get data files in data_path
    data_fn_list = [fn for fn in os.listdir(data_path) if fn.endswith(file_ext)]

    # generate xlsx file to save combined results to
    wb_sum = Workbook()

    # make summary sheet
    ws_comb_sum = wb_sum.active
    ws_comb_sum.title = "Summary"

    # make events sheet
    ws_comb_events = wb_sum.create_sheet("Events")

    # loop over data files to read data and load values
    for ith, fn in enumerate(data_fn_list):
        # make sure that current file obeys file format (and, if so, get exp conditions)
        matches = pat.match(fn)
        if matches is None:
            print("Invalid filename: {}".format(fn))
            continue
        exp_conditions = matches.groupdict()

        # get conditions from data filename
        starve_period = int(exp_conditions["starvePeriod"])
        food_type = int(exp_conditions["foodType"])
        food_type_units = exp_conditions["foodTypeUnits"]
        if food_type_units == "M":
            food_type = 1000*food_type

        # load current xlsx file
        data_path_curr = os.path.join(data_path, fn)
        headers_curr, data_curr = my_load_xlsx(data_path_curr)

        # separate summary and events data
        headers_sheet_names = [h.split('--')[0] for h in headers_curr]
        headers_var_names = [h.split('--')[1] for h in headers_curr]
        summary_idx = [i for i, x in enumerate(headers_sheet_names) if (x == "Summary")]
        events_idx = [i for i, x in enumerate(headers_sheet_names) if (x == "Events")]

        # ------------------------------------------------------------
        # SUMMARY
        # if writing to summary sheet for the first time, add headers
        if ith == 0:
            sum_header_names = [headers_var_names[ith] for ith in summary_idx]
            sum_header_names.append("Starvation Period (hr)")  # add fields for starvation period and  food type
            sum_header_names.append("Food Type (mM)")
            ws_comb_sum.append(sum_header_names)

        # get current summary data
        sum_data_curr = [data_curr[ith] for ith in summary_idx]

        # add starvation period and food type to list
        starve_period_list = len(sum_data_curr[0])*[starve_period]
        food_type_list = len(sum_data_curr[0])*[food_type]

        sum_data_curr.append(starve_period_list)
        sum_data_curr.append(food_type_list)

        # write current summary data to combined file
        for row_num in np.arange(len(sum_data_curr[0])):
            row_curr = [d[row_num] for d in sum_data_curr]
            ws_comb_sum.append(row_curr)

        # ------------------------------------------------------------
        # EVENTS
        # if writing to events sheet for the first time, add headers
        if ith == 0:
            events_header_names = [headers_var_names[ith] for ith in events_idx]
            events_header_names.append("Starvation Period (hr)")  # add fields for starvation period and  food type
            events_header_names.append("Food Type (mM)")
            ws_comb_events.append(events_header_names)

        # get current events data
        events_data_curr = [data_curr[ith] for ith in events_idx]

        # add starvation period and food type to list
        starve_period_list = len(events_data_curr[0]) * [starve_period]
        food_type_list = len(events_data_curr[0]) * [food_type]

        events_data_curr.append(starve_period_list)
        events_data_curr.append(food_type_list)

        # write current summary data to combined file
        for row_num in np.arange(len(events_data_curr[0])):
            row_curr = [d[row_num] for d in events_data_curr]
            ws_comb_events.append(row_curr)

        print("Completed data file {fname}".format(fname=fn))

    # SAVE XLSX FILE
    wb_sum.save(os.path.join(data_path, save_fn))
