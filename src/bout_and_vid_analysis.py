# -*- coding: utf-8 -*-
"""
Created on Wed Dec 12 09:00:55 2018

@author: Fruit Flies

Tools for combining tracking and feeding data from Visual Expresso expts
"""
# ------------------------------------------------------------------------------
import os
import sys

import numpy as np
from matplotlib import pyplot as plt
from matplotlib import cm

import re
import h5py

from load_hdf5_data import load_hdf5, my_add_h5_dset, my_add_dset_to_dict
from v_expresso_utils import idx_by_thresh
from bout_analysis_func import (check_data_set, plot_channel_bouts, bout_analysis, changepts_to_bouts, merge_meal_bouts,
                                check_bouts)
from v_expresso_gui_params import (analysisParams, trackingParams)
from v_expresso_image_lib import (visual_expresso_tracking_main, process_visual_expresso,
                                  hdf5_to_flyTrackData, save_vid_time_series,
                                  save_vid_summary, batch_plot_cum_dist,
                                  batch_plot_heatmap)

# from PIL import ImageTk, Image
import ast
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# -----------------------------------------------
# define some useful dictionaries/lists
summary_heading = ['Filename', 'Bank', 'Channel', 'Number of Meals',
                   'Total Volume (nL)', 'Total Duration Eating (s)',
                   'Latency to Eat (s)', 'Cumulative Dist. (cm)',
                   'Average Speed (cm/s)', 'Fraction Time Moving',
                   'Pre Meal Dist. (cm)', 'Food Zone Frac. (pre meal)',
                   'Food Zone Frac. (post meal)', 'Time (s)',
                   'Cumulative Dist. Time Series (cm)', 'Radial Dist. Time Series (cm)']

events_heading = ['Filename', 'Bank', 'Channel', 'Meal Number',
                  'Mealwise Start Time (s)', 'Mealwise End Time (s)',
                  'Mealwise Duration (s)', 'Mealwise Volume (nL)',
                  'Mealwise Dwell Time (s)',
                  'Mealwise Dwell Time Censoring (bool)']

summary_heading_dict = {'Filename': 'f_name',
                        'Bank': 'xp_name',
                        'Channel': 'ch_name',
                        'Number of Meals': 'num_meals',
                        'Total Volume (nL)': 'tot_vol',
                        'Total Duration Eating (s)': 'dur_eating',
                        'Latency to Eat (s)': 'eat_latency',
                        'Cumulative Dist. (cm)': 'cum_dist_max',
                        'Average Speed (cm/s)': 'avg_vel_mag',
                        'Fraction Time Moving': 'frac_moving',
                        'Pre Meal Dist. (cm)': 'pre_meal_dist',
                        'Food Zone Frac. (pre meal)': 'fz_frac_pre',
                        'Food Zone Frac. (post meal)': 'fz_frac_post',
                        'Time (s)': 't',
                        'Cumulative Dist. Time Series (cm)': 'cum_dist_ts',
                        'Radial Dist. Time Series (cm)': 'dist_mag_ts'}

# variables to put in events page
events_heading_dict = {'Filename': 'f_name',
                       'Bank': 'xp_name',
                       'Channel': 'ch_name',
                       'Meal Number': 'meal_num',
                       'Mealwise Start Time (s)': 'start_time',
                       'Mealwise End Time (s)': 'end_time',
                       'Mealwise Duration (s)': 'meal_dur',
                       'Mealwise Volume (nL)': 'meal_vol',
                       'Mealwise Dwell Time (s)': 'dwell_time',
                       'Mealwise Dwell Time Censoring (bool)': 'dwell_censor'}


# ------------------------------------------------------------------------------

# ------------------------------------------------------------------------------
# function to convert channel entries to basic filepath string for comparison
def channel2basic(channel_entry):
    filepath, bank_name, channel_name = channel_entry.split(', ', 2)
    filepath_no_ext = os.path.splitext(filepath)[0]
    basic_entry = filepath_no_ext + '_' + bank_name + '_' + channel_name

    return basic_entry


# ------------------------------------------------------------------------------
# function to convert video entries to basic filepath string for comparison
def vid2basic(vid_entry):
    basic_entry = os.path.splitext(vid_entry)[0]
    return basic_entry


# ------------------------------------------------------------------------------
# function to convert basic entries back to channel listbox format
def basic2channel(basic_entry):
    ent_split = basic_entry.split('_')
    channel_curr = '_'.join(ent_split[-2:])
    bank_curr = ent_split[-3]
    filepath_no_ext = '_'.join(ent_split[:-3])
    filepath = filepath_no_ext + '.hdf5'

    channel_entry = ', '.join([filepath, bank_curr, channel_curr])
    if sys.version_info[0] < 3:
        channel_entry = unicode(channel_entry)
    else:
        channel_entry = str(channel_entry)

    return channel_entry


# ------------------------------------------------------------------------------
# function to convert video entries to basic filepath string for comparison
def basic2vid(basic_entry, file_ext='.avi'):
    vid_entry = basic_entry + file_ext
    return vid_entry


# ------------------------------------------------------------------------------
# function that will take a list of basic entries and group them by expt
def group_expts_basic(basic_entry_list):
    dir_list = [os.path.dirname(ent) for ent in basic_entry_list]
    dir_list_unique = list(set(dir_list))

    expt_idx = np.zeros(len(dir_list), dtype=int)
    N_expt = []
    for ith, dr in enumerate(dir_list_unique):
        idx = [i for i, j in enumerate(dir_list) if j == dr]
        N_expt.append(len(idx))
        expt_idx[idx] = ith

    return expt_idx, N_expt


# ------------------------------------------------------------------------------
# function to check location and velocity of fly during putative meal bout
# CURRENTLY CHECKING:
#   - that fly gets sufficiently close to tip (min_dist_check)
#   - fly doesn't exceed a threshold velocity during meal (max_vel_check)
# TO ADD:
#   - make sure fly doesn't move too far from tip?
#   - use moving_ind?
def check_bout_wTracking(bouts, channel_t, vid_t, xcm, ycm, vel_mag, bout_params=analysisParams):
    # -----------------
    # read out params
    # MOVE_FRAC_THRESH = bout_params['feeding_move_frac_thresh']
    # MAX_DIST_THRESH = bout_params['feeding_dist_max']
    # MIN_DIST_THRESH = bout_params['feeding_dist_min']
    MIN_DIST_THRESH_X = bout_params['feeding_dist_min_x']
    MIN_DIST_THRESH_Y = bout_params['feeding_dist_min_y']
    MAX_VEL_THRESH = bout_params['feeding_vel_max']
    # dist_max_prctile_lvl = 90  # 75
    dist_min_prctile_lvl = 50  # 25
    vel_max_prctile_lvl = 50  # 50

    # ---------------------------
    # get number of bouts
    N_bouts = bouts.shape[1]

    # ----------------------------------------------
    # loop over bouts and check all heuristics
    bout_check = np.zeros(N_bouts)  # intialize indices
    for ith in np.arange(N_bouts):
        # first need to match up timing between video and channel
        vid_start_idx = np.argmin(np.abs(vid_t - channel_t[bouts[0, ith]]))
        vid_end_idx = np.argmin(np.abs(vid_t - channel_t[bouts[1, ith]]))
        vid_bout_dur = float(vid_end_idx - vid_start_idx)

        # check that the fly doesn't exceed a certain velocity
        vel_mag_bout = vel_mag[vid_start_idx:vid_end_idx]
        vel_max_prctile = np.percentile(vel_mag_bout, vel_max_prctile_lvl)
        max_vel_check = (vel_max_prctile < MAX_VEL_THRESH)

        # check that fly is close to cap tip
        x_bout = np.abs(xcm[vid_start_idx:vid_end_idx])
        y_bout = np.abs(ycm[vid_start_idx:vid_end_idx])
        dist_min_prctile_x = np.percentile(x_bout, dist_min_prctile_lvl)
        dist_min_prctile_y = np.percentile(y_bout, dist_min_prctile_lvl)
        min_dist_check = (dist_min_prctile_x < MIN_DIST_THRESH_X) & (dist_min_prctile_y < MIN_DIST_THRESH_Y)

        # check if all conditions are met
        bout_check[ith] = min_dist_check & max_vel_check  # moving_chk & #max_dist_check &

    return bout_check


# UNUSED:
# # check that the fly is not moving for a certain fraction of bout
# total_moving_bout = float(np.sum(moving_ind[vid_start_idx:vid_end_idx]))
# moving_chk = total_moving_bout / vid_bout_dur
# moving_chk = (moving_chk < MOVE_FRAC_THRESH)
#
# dist_mag_bout = dist_mag[vid_start_idx:vid_end_idx]
# dist_max_prctile = np.percentile(dist_mag_bout,dist_max_prctile_lvl)
# dist_min_prctile = np.percentile(dist_mag_bout,dist_min_prctile_lvl)
# max_dist_check = (dist_max_prctile < MAX_DIST_THRESH)

# ------------------------------------------------------------------------------
# function to check bouts based on tracking data
def bout_analysis_wTracking(filename, bank_name, channel_name, bts=[], vols=[], time=[], dset_sm=[], bt_chgpts=[],
                            bout_params=analysisParams, saveFlag=False, plotFlag=False, debugBoutFlag=False,
                            debugTrackingFlag=False, split_meals_flag=True, runTrackingFlag=False):
    # --------------------------------
    # load channel data and get bouts
    # --------------------------------
    if (np.asarray(bts).size > 0) and (np.asarray(vols).size > 0) and (np.asarray(bt_chgpts).size > 0):
        bouts = bts
        volumes = vols
        dset_smooth = dset_sm
        t = time
        bout_changepts = bt_chgpts
    else:
        dset, t = load_hdf5(filename, bank_name, channel_name)
        bad_data_flag, dset, t, frames = check_data_set(dset, t)

        if not bad_data_flag:
            dset_smooth, bouts, volumes, bout_changepts = bout_analysis(dset, frames, debug_mode=debugBoutFlag)
        else:
            print('Problem loading data set')
            dset_smooth = np.nan
            bouts_corrected = np.nan
            volumes_corrected = np.nan
            return dset_smooth, bouts_corrected, volumes_corrected

    # --------------------------------
    # load tracking data
    # --------------------------------
    file_path, filename_hdf5 = os.path.split(filename)
    data_prefix = os.path.splitext(filename_hdf5)[0]
    filename_vid = data_prefix + '_' + bank_name + "_" + channel_name + '.avi'
    filename_vid_analysis = data_prefix + '_' + bank_name + "_" + channel_name + '_TRACKING_PROCESSED.hdf5'
    filename_vid_track = data_prefix + '_' + bank_name + "_" + channel_name + '_TRACKING.hdf5'

    filename_vid_full = os.path.join(file_path, filename_vid)
    filename_vid_full = os.path.abspath(filename_vid_full)

    filename_vid_analysis_full = os.path.join(file_path, filename_vid_analysis)
    filename_vid_analysis_full = os.path.abspath(filename_vid_analysis_full)

    if os.path.exists(filename_vid_analysis_full):
        flyTrackData = hdf5_to_flyTrackData(file_path, filename_vid_analysis)
    elif os.path.exists(filename_vid_full) and runTrackingFlag:
        print('{} has not yet been analyzed. Doing so now...'.format(filename_vid_full))
        visual_expresso_tracking_main(file_path, filename_vid, DEBUG_BG_FLAG=debugTrackingFlag,
                                      DEBUG_CM_FLAG=debugTrackingFlag, SAVE_DATA_FLAG=True, ELLIPSE_FIT_FLAG=False,
                                      PARAMS=trackingParams)
        flyTrackData = process_visual_expresso(file_path, filename_vid_track, SAVE_DATA_FLAG=True, DEBUG_FLAG=False)
    else:
        print('Error: no matching video file found. check directory')
        bouts_corrected = bouts
        volumes_corrected = volumes
        return dset_smooth, bouts_corrected, volumes_corrected

    # -------------------------------------------------
    # get fly position/velocity during meal bouts
    # -------------------------------------------------
    channel_t = t
    vid_t = flyTrackData['t']
    xcm_smooth = flyTrackData['xcm_smooth']
    ycm_smooth = flyTrackData['ycm_smooth']
    dist_mag = np.sqrt(xcm_smooth ** 2 + ycm_smooth ** 2)
    vel_mag = flyTrackData['vel_mag']

    # -----------------------------------------------
    # check meal bouts using tracking data
    # -----------------------------------------------

    # switch procedure depending on whether or not we're splitting meals prior to checking them
    if split_meals_flag:
        # get array of bout changepoints -- allows finer grain checking of each meal bout
        changept_meal_bouts = changepts_to_bouts(bout_changepts)

        # apply heuristics to check each CHANGEPOINT meal bout
        chgpt_bout_check = check_bout_wTracking(changept_meal_bouts, channel_t, vid_t, xcm_smooth, ycm_smooth, vel_mag,
                                          bout_params=bout_params)

        # now remove the bouts that fail to meet distance/moving criteria
        chgpt_bout_check = np.asarray(chgpt_bout_check, dtype=int)
        chgpt_bouts_corrected = changept_meal_bouts[:, (chgpt_bout_check == 1)]

        # merge changepoint bouts to get normal bout list
        bouts_corrected = merge_meal_bouts(chgpt_bouts_corrected)

        # check to make sure that, in the process of splitting/merging, we didn't violate feeding bout heuristics
        _, bouts_corrected = check_bouts(bouts_corrected, dset_smooth, analysis_params=bout_params)

        # get volumes for these new, corrected meal bouts
        volumes_corrected = dset_smooth[bouts_corrected[0, :]] - dset_smooth[bouts_corrected[1, :]]
    else:
        # just check full (merged) meals
        bout_check = check_bout_wTracking(bouts, channel_t, vid_t, xcm_smooth, ycm_smooth, vel_mag,
                                          bout_params=bout_params)

        # convert logical index array to int and use to get good meals/volumes (not sure why i did the int conversion?)
        bout_check = np.asarray(bout_check, dtype=int)
        bouts_corrected = bouts[:, (bout_check == 1)]
        volumes_corrected = volumes[(bout_check == 1)]

    # ---------------------------------------------------
    # plot results?
    if plotFlag:
        fig_cor, ax1_cor, ax2_cor = plot_channel_bouts(dset, dset_smooth, t, bouts_corrected)
        fig, ax1, ax2 = plot_channel_bouts(dset, dset_smooth, t, bouts)
        fig_cor.suptitle('With tracking')
        fig.suptitle('No tracking')
        fig.show()

    # in addition to standard bout debugging plots, create plot that shows
    #   location and velocity during putative bout
    if debugBoutFlag:
        fig, ax1, ax2 = plot_channel_bouts(dset, dset_smooth, t, bouts_corrected)
        color = 'green'

        # plot distance from tip on top of raw data        
        ax1_twin = ax1.twinx()
        ax1_twin.set_ylabel('Dist. from tip [cm]', color=color)
        ax1_twin.plot(vid_t, dist_mag, color=color)
        ax1_twin.tick_params(axis='y', labelcolor=color)
        ax1_twin.set_xlim([0, np.max(vid_t)])

        # plot speed on top of smooth data
        ax2_twin = ax2.twinx()
        ax2_twin.set_ylabel('Speed [cm/s]', color=color)
        ax2_twin.plot(vid_t, flyTrackData['vel_mag'], color=color)
        ax2_twin.tick_params(axis='y', labelcolor=color)
        ax2_twin.set_xlim([0, np.max(vid_t)])

        fig.tight_layout()  # otherwise the right y-label is slightly clipped
        fig.show()
    # save results?
    if saveFlag:
        print('under construction')

    return dset_smooth, bouts_corrected, volumes_corrected


# ------------------------------------------------------------------------------
# function to create dictionary with both tracking and feeding data
def merge_v_expresso_data(dset, dset_smooth, channel_t, frames, bouts, volumes, flyTrackData, boutParams=analysisParams):
    flyCombinedData = dict()
    for key, value in flyTrackData.items():
        flyCombinedData[key] = value

    filename = flyCombinedData['filename']
    filename_split = filename.split('_')
    if filename_split[-2] == 'TRACKING':
        filename_new = '_'.join(filename_split[:-2]) + '_COMBINED_DATA.hdf5'
    else:
        filename_new = '_'.join(filename_split[:-1]) + '_COMBINED_DATA.hdf5'

    flyCombinedData.update({'channel_dset': dset,
                            'channel_dset_units': 'nL',
                            'channel_dset_long_name': 'Liquid Level (nL)',
                            'channel_dset_smooth': dset_smooth,
                            'channel_dset_smooth_units': 'nL',
                            'channel_dset_smooth_long_name': 'Liquid Level (nL)',
                            'channel_t': channel_t,
                            'channel_t_units': 's',
                            'channel_t_long_name': 'Time (s)',
                            'channel_frames': frames,
                            'bouts': bouts,
                            'bouts_units': 'idx',
                            'bouts_long_name': 'Feeding Bout Idx',
                            'volumes': volumes,
                            'volumes_units': 'nL',
                            'volumes_long_name': 'Meal Volume (nL)',
                            'boutParams': boutParams,
                            'filename': filename_new})

    return flyCombinedData


# ------------------------------------------------------------------------------
# function to save flyCombinedData dict to hdf5 file
def flyCombinedData_to_hdf5(flyCombinedData):
    filename = flyCombinedData['filename']
    filepath = flyCombinedData['filepath']

    save_filename = os.path.join(filepath, filename)

    with h5py.File(save_filename, 'w') as f:
        # ----------------------------------
        # time and parameters
        # ----------------------------------
        # video time
        my_add_h5_dset(f, 'Time', 't', flyCombinedData['t'], units='s', long_name='Time (s)')
        # channel time
        my_add_h5_dset(f, 'Time', 'channel_t', flyCombinedData['channel_t'], units='s', long_name='Time (s)')
        # pixel to centimeter conversion factor
        my_add_h5_dset(f, 'Params', 'pix2cm', flyCombinedData['PIX2CM'], units='pix/cm', long_name='Pixels per cm')
        # tracking params
        if 'trackingParams' in flyCombinedData:
            my_add_h5_dset(f, 'Params', 'trackingParams', str(flyCombinedData['trackingParams']))
        # bout params
        if 'boutParams' in flyCombinedData:
            my_add_h5_dset(f, 'Params', 'boutParams', str(flyCombinedData['boutParams']))

        # ----------------------------------
        # body velocity
        # ----------------------------------
        my_add_h5_dset(f, 'BodyVel', 'vel_x', flyCombinedData['xcm_vel'], units='cm/s', long_name='X Velocity (cm/s)')
        my_add_h5_dset(f, 'BodyVel', 'vel_y', flyCombinedData['ycm_vel'], units='cm/s', long_name='Y Velocity (cm/s)')
        my_add_h5_dset(f, 'BodyVel', 'vel_mag', flyCombinedData['vel_mag'], units='cm/s', long_name='Speed (cm/s)')
        my_add_h5_dset(f, 'BodyVel', 'moving_ind', flyCombinedData['moving_ind'], units='idx', long_name='Moving Index')

        # ----------------------------------
        # body position and image info
        # ----------------------------------
        my_add_h5_dset(f, 'BodyCM', 'xcm_smooth', flyCombinedData['xcm_smooth'], units='cm',
                       long_name='X Position (cm)')
        my_add_h5_dset(f, 'BodyCM', 'ycm_smooth', flyCombinedData['ycm_smooth'], units='cm',
                       long_name='Y Position (cm)')
        my_add_h5_dset(f, 'BodyCM', 'cum_dist', flyCombinedData['cum_dist'], units='cm',
                       long_name='Cumulative Dist. (cm)')

        if 'interp_idx' in flyCombinedData:
            try:
                my_add_h5_dset(f, 'BodyCM', 'interp_idx', flyCombinedData['interp_idx'], units='idx')
            except TypeError:
                pass

        # unprocessed data/information from tracking output file
        my_add_h5_dset(f, 'BodyCM', 'xcm', flyCombinedData['xcm'], units='cm', long_name='Raw X Position (cm)')
        my_add_h5_dset(f, 'BodyCM', 'ycm', flyCombinedData['ycm'], units='cm', long_name='Raw Y Position (cm)')
        my_add_h5_dset(f, 'ROI', 'roi', flyCombinedData['ROI'], units='pix', long_name='Region of Interest (pix)')
        my_add_h5_dset(f, 'BG', 'bg', flyCombinedData['BG'])
        my_add_h5_dset(f, 'CAP_TIP', 'cap_tip', flyCombinedData['cap_tip'], units='pix')
        my_add_h5_dset(f, 'CAP_TIP', 'cap_tip_orientation', flyCombinedData['cap_tip_orientation'])

        # body angle (?)
        if 'body_angle' in flyCombinedData:
            try:
                my_add_h5_dset(f, 'BodyAngle', 'body_angle', flyCombinedData['body_angle'], units='deg',
                               long_name='Body Angle (deg)')
            except TypeError:
                pass

        # ----------------------------------
        # feeding data
        # ----------------------------------
        my_add_h5_dset(f, 'Feeding', 'dset_raw', flyCombinedData['channel_dset'], units='nL',
                       long_name='Liquid Level (nL)')
        my_add_h5_dset(f, 'Feeding', 'dset_smooth', flyCombinedData['channel_dset_smooth'], units='nL',
                       long_name='Liquid Level (nL)')
        my_add_h5_dset(f, 'Feeding', 'bouts', flyCombinedData['bouts'], units='idx', long_name='Feeding Bout Idx')
        my_add_h5_dset(f, 'Feeding', 'volumes', flyCombinedData['volumes'], units='nL', long_name='Meal Volume (nL)')


# ------------------------------------------------------------------------------
# function to load hdf5 file as flyCombinedData dict
def hdf5_to_flyCombinedData(filepath, filename):
    # file information    
    flyCombinedData = {'filepath': filepath,
                       'filename': filename}

    data_prefix = os.path.splitext(filename)[0]
    xp_name = data_prefix.split('_')[-5]
    channel_name = 'channel_' + data_prefix.split('_')[-3]

    # bank and channel names
    flyCombinedData['xp_name'] = xp_name
    flyCombinedData['channel_name'] = channel_name

    # tracking results
    filename_full = os.path.join(filepath, filename)
    filename_full = os.path.abspath(filename_full)
    with h5py.File(filename_full, 'r') as f:

        # --------------------------
        # Image info
        # --------------------------
        try:
            flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'ROI', f, 'ROI', 'roi')
            flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'cap_tip', f, 'CAP_TIP', 'cap_tip')
            flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'cap_tip_orientation', f, 'CAP_TIP',
                                                  'cap_tip_orientation', scalar_flag=True)
            flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'BG', f, 'BG', 'bg')
        except KeyError:
            flyCombinedData = {}
            return flyCombinedData


        # --------------------------
        # params
        # --------------------------
        flyCombinedData['PIX2CM'] = f['Params']['pix2cm'][()]
        if (sys.version_info[0] < 3):
            flyCombinedData['trackingParams'] = ast.literal_eval(f['Params']['trackingParams'][:])
            flyCombinedData['boutParams'] = ast.literal_eval(f['Params']['boutParams'][:])

        # --------------------------
        # Time
        # --------------------------
        t = f['Time']['t'][:]
        channel_t = f['Time']['channel_t'][:]

        flyCombinedData['frames'] = np.arange(len(t))
        # flyCombinedData['t'] = t
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 't', f, 'Time', 't')
        flyCombinedData['channel_frames'] = np.arange(len(channel_t))
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'channel_t', f, 'Time', 'channel_t')

        # --------------------------
        # Trajectory info
        # --------------------------
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'xcm', f, 'BodyCM', 'xcm')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'ycm', f, 'BodyCM', 'ycm')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'xcm_smooth', f, 'BodyCM', 'xcm_smooth')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'ycm_smooth', f, 'BodyCM', 'ycm_smooth')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'cum_dist', f, 'BodyCM', 'cum_dist')

        try:
            flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'interp_idx', f, 'BodyCM', 'interp_idx')
        except (KeyError, ValueError):
            flyCombinedData['interp_idx'] = np.nan

        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'xcm_vel', f, 'BodyVel', 'vel_x')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'ycm_vel', f, 'BodyVel', 'vel_y')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'vel_mag', f, 'BodyVel', 'vel_mag')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'moving_ind', f, 'BodyVel', 'moving_ind')

        try:
            flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'body_angle', f, 'BodyAngle', 'body_angle')
        except (KeyError, ValueError):
            flyCombinedData['body_angle'] = np.nan

        # --------------------------
        # Feeding data
        # --------------------------
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'channel_dset', f, 'Feeding', 'dset_raw')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'channel_dset_smooth', f, 'Feeding', 'dset_smooth')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'bouts', f, 'Feeding', 'bouts')
        flyCombinedData = my_add_dset_to_dict(flyCombinedData, 'volumes', f, 'Feeding', 'volumes')

    return flyCombinedData


# ------------------------------------------------------------------------------
# function to match up expresso channel and camera timing
def interp_channel_time(flyCombinedData):
    cam_t = flyCombinedData['t']
    channel_t = flyCombinedData['channel_t']
    dset_raw = flyCombinedData['channel_dset']
    dset_smooth = flyCombinedData['channel_dset_smooth']
    bouts = flyCombinedData['bouts']

    if bouts.size > 1:
        bouts_start_t = channel_t[bouts[0, :]]
        bouts_end_t = channel_t[bouts[1, :]]

        bouts_idx1_cam = np.zeros(bouts_start_t.shape, dtype=int)
        bouts_idx2_cam = np.zeros(bouts_end_t.shape, dtype=int)

        for ith in np.arange(len(bouts_idx1_cam)):
            bouts_idx1_cam[ith] = np.argmin(np.abs(bouts_start_t[ith] - cam_t))
            bouts_idx2_cam[ith] = np.argmin(np.abs(bouts_end_t[ith] - cam_t))

        bouts_cam = np.vstack((bouts_idx1_cam, bouts_idx2_cam))
    else:
        bouts_cam = np.empty((2, 0), dtype=int)

    dset_cam = np.interp(cam_t, channel_t, dset_raw)
    dset_smooth_cam = np.interp(cam_t, channel_t, dset_smooth)

    return dset_cam, dset_smooth_cam, bouts_cam


# ------------------------------------------------------------------------------
# function to calculate dwell time after each meal
def get_dwell_time(bouts, channel_t, dist_mag, vid_t, fz_rad=analysisParams['food_zone_rad']):
    try:
        N_meals = bouts.shape[1]
    except IndexError:
        N_meals = 0

    dwell_times = np.full((N_meals, 1), np.nan)
    censoring = np.full((N_meals, 1), 0)
    for meal_num in range(N_meals):
        # find ending of current meal
        meal_end_t = channel_t[bouts[1, meal_num]]

        # find first time after meal end that fly moves away from cap tip
        post_meal_idx = np.where((vid_t > meal_end_t) & (dist_mag > fz_rad))[0]

        if (post_meal_idx.size == 0):
            dwell_times[meal_num] = np.max(vid_t) - meal_end_t
            censoring[meal_num] = 1
        else:
            leave_idx = post_meal_idx[0]
            dwell_times[meal_num] = vid_t[leave_idx] - meal_end_t
            censoring[meal_num] = 0

    return dwell_times, censoring


# ------------------------------------------------------------------------------
# function to get data aligned to a meal
def get_meal_aligned_data(h5_fn, var, window_left_sec=0, window_right_sec=10, meal_num=0):
    # initialize output
    data_aligned = None

    # load data from hdf5 file
    filepath, filename = os.path.split(h5_fn)
    flyCombinedData = hdf5_to_flyCombinedData(filepath, filename)

    # check if meal data exists in file. if so, read out and align to video time. otherwise return
    if 'bouts' in flyCombinedData:
        _, _, bouts_cam = interp_channel_time(flyCombinedData)
    else:
        print('No feeding data in {} -- skipping'.format(filename))
        return

    # check that this data file contains the specified meal (i.e. if we're looking for data aligned to 4th meal, check
    # make sure that there are 4 meals in data file
    # print(bouts_cam.shape)
    N_meals = bouts_cam.shape[1]
    if N_meals <= meal_num:  # or (N_meals < 1):
        return

    # convert time window (given in seconds) into index
    t = flyCombinedData['t']
    dt = np.nanmean(np.diff(t))
    window_left_idx = round(window_left_sec / dt)
    window_right_idx = round(window_right_sec / dt)

    # get indices for time window around meal (distance from meal end time is specified by window_left and window_right)
    meal_end = bouts_cam[1, meal_num]
    idx1 = np.max([meal_end - window_left_idx, 0])
    idx2 = np.min([meal_end + window_right_idx, len(t) - 1])

    # ----------------------------
    # LOAD DATA
    # ----------------------------
    # check if selected var data exists in file. for most cases, var should be a key in the flyCombinedData dict. if
    # it's one of a few other data types, we need to define it here. if neither are found, return
    if var in flyCombinedData:
        data_var = flyCombinedData[var]

        # get data for specified indices (window around meal)
        data_aligned = data_var[idx1:idx2]

    elif var == 'dist_mag':
        # load x and y cm position data
        xdata = flyCombinedData['xcm_smooth']
        ydata = flyCombinedData['ycm_smooth']

        # get data for specified indices (window around meal) and center by subtracting off value at meal end
        xdata_aligned = xdata[idx1:idx2] - xdata[meal_end]
        ydata_aligned = ydata[idx1:idx2] - ydata[meal_end]

        # get radial distance by taking root sum of x,y
        data_aligned = np.sqrt(xdata_aligned**2 + ydata_aligned**2)
    else:
        print('No {} data in {} -- skipping'.format(var, filename))
        return

    # for some variables, want to subtract off value at beginning of array to align
    init_sub_vars = ['t', 'xcm_smooth', 'ycm_smooth']
    if var in init_sub_vars:
        data_aligned = data_aligned - data_var[meal_end]

    # return meal aligned data
    return data_aligned

# ---------------------------------------------------------------------------------------------
# function to plot feeding-bout-end aligned data (old -- tries to group data, need to update)
def plot_bout_aligned_var(basic_entries, varx='xcm_smooth', vary='ycm_smooth', window_left_sec=0, window_right_sec=300,
                          meal_num=0, figsize=(6, 6), save_flag=False, save_filename=None, varx_name=None,
                          vary_name=None, one_x_column_flag=False):
    # suffix for filename with both feeding and tracking data
    data_suffix = '_COMBINED_DATA.hdf5'

    # define colormap for plot
    # colors = plt.cm.Set1(range(len(basic_entries)))
    plt.rcParams["axes.prop_cycle"] = plt.cycler("color", plt.cm.Set1.colors)

    # ---------------------------------------------------------------
    # initialize save output
    # ---------------------------------------------------------------
    # for saving options, first check that we have a valid save name
    if save_flag and not save_filename:
        save_flag = False

    # but, if we do have a valid savename, initialize storage for all data
    if save_flag:
        data_all = []
        has_data_flag = []
    # --------------------------------------------
    # initialize figure and axis
    # fig, ax = plt.subplots(1, 1, figsize=figsize)
    fig, ax = plt.subplots(1, 1)

    # ----------------------------------------------------
    # loop over data files
    for ith, ent in enumerate(basic_entries):
        # get full filename for current hdf5 analysis file, as well as file id
        hdf5_filename = ent + data_suffix
        if os.path.exists(hdf5_filename):
            ent_id = os.path.basename(ent)
            # read out meal aligned varx and vary data
            data_x = get_meal_aligned_data(hdf5_filename, varx, window_left_sec=window_left_sec,
                                           window_right_sec=window_right_sec, meal_num=meal_num)
            data_y = get_meal_aligned_data(hdf5_filename, vary, window_left_sec=window_left_sec,
                                           window_right_sec=window_right_sec, meal_num=meal_num)
        else:
            data_x = None
            data_y = None

        # make sure we got data for both x and y
        if data_x is None or data_y is None:
            print('Failed to load data for {} -- skipping'.format(hdf5_filename))
            if save_flag:
                # if saving data, need to keep track of empty array elements
                has_data_flag.append(False)
            continue

        # plot data on current axes
        ax.plot(data_x, data_y, '-', label=ent_id, markersize=2, linewidth=0.75)

        # if we're saving data, add current x,y data to "all" list
        if save_flag:
            # NB: appending x then y in this order so we can have an x and y column for each file
            data_all.append(data_x)
            data_all.append(data_y)
            has_data_flag.append(True)
    # --------------------------------------------------------------------
    # axis properties (once plotting has finished)
    plt.legend(loc='upper left', bbox_to_anchor=(1, 1.05), fontsize='x-small')

    # show figure
    fig.show()

    # set axis labels, if they exist
    if varx_name:
        ax.set_xlabel(varx_name)
    if vary_name:
        ax.set_ylabel(vary_name)

    # make plot window tight
    fig.tight_layout()

    # --------------------------------------------------------------------
    # if saving, write combined data to workbook
    if save_flag:
        # initialize workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = vary + " vs " + varx

        # write meal number in first row
        meal_num_heading = ['Meal Number:', str(meal_num + 1)]  # NB: adding one to account for Python indexing
        ws.append(meal_num_heading)

        # remove files in which there is no data (valid entries are ones with data)
        # basic_entries_valid = [ent for (kth, ent) in enumerate(basic_entries) if has_data_flag[kth]]
        basic_entries_valid = [ent for (tf, ent) in zip(has_data_flag, basic_entries) if tf]

        # from here, switch how we write to file depending on whether or not we want 2 columns per file, or a single x
        # data column and then a 1 y data column per file
        if one_x_column_flag:
            # -----------------------------------------------------------------------------
            # is this case, each file gets a "y" column, and there's one global "x" column
            # -----------------------------------------------------------------------------
            # write filenames in second row (with a spacer at the beginning for "x" column)
            fn_heading = []
            fn_heading.append(' ')
            for ent in basic_entries_valid:
                fn_heading.append(os.path.basename(ent))
            ws.append(fn_heading)

            # write variable names in third row
            if varx_name and vary_name:
                var_heading = [varx_name] + len(basic_entries_valid) * [vary_name]
            else:
                var_heading = [varx] + len(basic_entries_valid) * [vary]
            ws.append(var_heading)

            # convert list to numpy array and take transpose so we have N x (M+1) matrix, where N = # data points, M = # files
            data_all = [data_all[0]] + data_all[1::2]
            data_all = np.transpose(np.vstack(data_all))
        else:
            # -----------------------------------------------------------------------------
            # in this case, each file gets a "x" and "y" column
            # -----------------------------------------------------------------------------
            # write filenames in second row (with a spacer in between each)
            fn_heading = []
            for ent in basic_entries_valid:
                fn_heading.append(os.path.basename(ent))
                fn_heading.append(' ')
            ws.append(fn_heading)

            # write variable names in third row
            if varx_name and vary_name:
                var_heading = len(basic_entries_valid) * [varx_name, vary_name]
            else:
                var_heading = len(basic_entries_valid) * [varx, vary]
            ws.append(var_heading)

            # convert list to numpy array and take transpose so we have N x 2M matrix, where N = # data points, M = # files
            data_all = np.transpose(np.vstack(data_all))

        # ---------------------------------------------------
        # the following should work for either column format
        # ---------------------------------------------------
        # loop over rows and write to file
        for row in data_all:
            ws.append(list(row))

        # set column width to be wider (more readable)
        for col in ws.columns:
            col_str = get_column_letter(col[0].column)
            ws.column_dimensions[col_str].width = 22
        # save workbook
        wb.save(save_filename)

    return fig, ax


# ---------------------------------------------------------------------------------------------
# function to plot feeding-bout-end aligned data (old -- tries to group data, need to update)
def plot_bout_aligned_var_old(basic_entries, var='vel_mag', window=300, N_meals=4, figsize=(6, 10), saveFlag=False):
    # suffix for filename with both feeding and tracking data
    data_suffix = '_COMBINED_DATA.hdf5'

    # group data files based on their parent directory
    expt_idx, N_expt = group_expts_basic(basic_entries)

    # get color maps for different experiment groupings
    cmap_name_list = ['Reds', 'Blues', 'Greens', 'Purples', 'Oranges', 'YlOrBr', 'PuBu', 'YlOrRd', 'BuGn']
    cmap_mat = np.array([])
    for jth in np.arange(len(N_expt)):
        cmap_curr = cm.get_cmap(cmap_name_list[jth])
        c_val = np.arange(N_expt[jth], dtype=float) + 1.0
        normalized_c_val = c_val / np.max(c_val)
        color_vecs = cmap_curr(tuple(normalized_c_val))
        if jth == 0:
            cmap_mat = (color_vecs)
        else:
            cmap_mat = np.vstack((cmap_mat, color_vecs))

    # create plot
    fig, ax_arr = plt.subplots(N_meals, figsize=figsize, sharex=True, sharey=True)

    # expt_idx_counter = expt_idx[0]
    # cc = 0
    for ith, ent in enumerate(basic_entries):
        # load combined data file
        filename_full = ent + data_suffix
        filepath, filename = os.path.split(filename_full)
        flyCombinedData = hdf5_to_flyCombinedData(filepath, filename)

        # current color for plots
        color_vec_curr = cmap_mat[ith, :]

        # assign data and select variable to plot
        t = flyCombinedData['t']
        t_plot = t[:window] - t[0]
        if var == 'vel_mag':
            var_curr = flyCombinedData[var]
        elif var == 'dist_mag':
            var_curr = np.sqrt(flyCombinedData['xcm_smooth'] ** 2 + flyCombinedData['ycm_smooth'] ** 2)
        else:
            print('do not have this option yet')
            return

        # since we're plotting tracking results, need bout index on cam time
        _, _, bouts_cam = interp_channel_time(flyCombinedData)

        # now loop through and plot
        N_meals_curr = np.min([bouts_cam.shape[1], N_meals])
        for kth in np.arange(N_meals_curr):
            idx1 = bouts_cam[1, kth]
            idx2 = np.min([bouts_cam[1, kth] + window, len(t) - 1])
            var_to_plot = var_curr[idx1:idx2]
            ax_arr[kth].plot(t_plot, var_to_plot, '-', color=color_vec_curr)

    for ax_num in np.arange(N_meals):
        if var == 'vel_mag':
            ax_arr[ax_num].set_ylabel('Speed (cm/s)')
        elif var == 'dist_mag':
            ax_arr[ax_num].set_ylabel('Distance from tip (cm)')
        else:
            print('do not have this option yet')
            return

        ax_arr[ax_num].set_title('Aligned to meal {} end'.format(ax_num + 1))
        if ax_num == (N_meals - 1):
            ax_arr[ax_num].set_xlabel('Time (s)')

    # show figure
    fig.show()

    return fig


# ------------------------------------------------------------------------------
# function to save time series of combined data to CSV format
def save_comb_time_series(data_filenames, savedir):
    print("Saving combined time series...")
    # loop through each data file
    for data_fn in data_filenames:
        if not os.path.exists(os.path.abspath(data_fn)):
            print(data_fn + ' not yet analyzed--failed to save')
        else:
            # load combined data file
            filepath, filename = os.path.split(data_fn)
            filename_noExt, _ = os.path.splitext(filename)
            flyCombinedData = hdf5_to_flyCombinedData(filepath, filename)

            # get save name for csv file
            csv_filename = os.path.join(savedir, filename_noExt + ".csv")
            column_headers = ['Frame', 'Time (s)', 'Channel Raw (nL)',
                              'Channel Smoothed (nL)', 'Feeding (bool)',
                              'X Position (cm)', 'Y Position (cm)',
                              'Cumulative Dist. (cm)', 'X Velocity (cm/s)',
                              'Y Velocity (cm/s)', 'Speed (cm/s)', 'Moving Idx']
            # ---------------------------------------------------------
            # load in relevant kinematics data
            t = flyCombinedData['t']
            frames = np.arange(0, t.size)

            xcm = flyCombinedData['xcm_smooth']
            ycm = flyCombinedData['ycm_smooth']
            cum_dist = flyCombinedData['cum_dist']

            x_vel = flyCombinedData['xcm_vel']
            y_vel = flyCombinedData['ycm_vel']
            vel_mag = flyCombinedData['vel_mag']
            moving_ind = flyCombinedData['moving_ind']

            # ---------------------------------------------------------
            # load in relevant bout data
            # (make sure that time values match up for bouts and tracking)
            dset_cam, dset_smooth_cam, bouts_cam = interp_channel_time(flyCombinedData)

            # generate array with same length as t that has "1" if fly is 
            # feeding and "0" otherwise
            feeding_boolean = np.zeros([1, dset_cam.size])
            for i in np.arange(bouts_cam.shape[1]):
                feeding_boolean[0, bouts_cam[0, i]:bouts_cam[1, i]] = 1

            # ---------------------------------------------------------
            # combine data into one matrix 
            row_mat = np.vstack((frames, t, dset_cam, dset_smooth_cam,
                                 feeding_boolean, xcm, ycm, cum_dist, x_vel,
                                 y_vel, vel_mag, moving_ind))
            row_mat = np.transpose(row_mat)

            # ------------------------------------------------------------------
            # generate csv file, write column headings, then write data in rows
            if sys.version_info[0] < 3:
                out_path = open(csv_filename, mode='wb')
            else:
                out_path = open(csv_filename, 'w', newline='')

            # csv writer object    
            save_writer = csv.writer(out_path)

            # write headings for filename + data column labels
            save_writer.writerow([filepath])
            save_writer.writerow(column_headers)

            # write data rows
            for row in row_mat:
                save_writer.writerow(row)

            out_path.close()
            print("Done saving {}".format(csv_filename))

    print("Completed saving time series data")


# ------------------------------------------------------------------------------
# function to save summary of combined data to XLSX format
def save_comb_summary(entry_list, xlsx_filename,
                      fz_rad=analysisParams['food_zone_rad']):
    print('Saving to {} ...'.format(xlsx_filename))
    # info for converting file types
    data_suffix = '_COMBINED_DATA.hdf5'

    # ------------------------------------------------------------------------
    # first check if the combined data exists -- if not, save whatever summary
    # we can (i.e. just tracking data)
    comb_data_idx = [os.path.exists(os.path.abspath(ent + data_suffix)) for
                     ent in entry_list]
    if not any(comb_data_idx):
        # check for tracking data
        track_suffix = '_TRACKING_PROCESSED.hdf5'
        track_data_idx = [os.path.exists(os.path.abspath(ent + track_suffix))
                          for ent in entry_list]

        # if we find tracking data, save just tracking summary
        if any(track_data_idx):
            print('No feeding data -- saving just tracking data')
            vid_suffix = '.avi'
            vid_filenames = [os.path.abspath(ent + vid_suffix) for ent in
                             entry_list]
            save_vid_summary(vid_filenames, xlsx_filename)
            return

        # if there's neither combined nor tracking data, we assume there's 
        # feeding (channel) data
        print('No tracking data -- saving just feeding data')

        # import necessary functions
        from batch_bout_analysis_func import (batch_bout_analysis,
                                              save_batch_xlsx)

        # convert list elements to be readable for batch channel analysis
        batch_list = [basic2channel(ent) for ent in entry_list]

        # plug in values for time limits/bin size
        tmin = np.nan
        tmax = np.nan
        tbin = 20

        # calculate batch channel stats
        (bouts_list, name_list, volumes_list, consumption_per_fly,
         duration_per_fly, latency_per_fly) = \
            batch_bout_analysis(batch_list, tmin, tmax, tbin,
                                plotFlag=False, combAnalysisFlag=False)

        # save batch channel results 
        save_batch_xlsx(xlsx_filename, bouts_list, name_list, volumes_list,
                        consumption_per_fly, duration_per_fly, latency_per_fly)

        print('Completed saving {}'.format(xlsx_filename))
        return
    elif not all(comb_data_idx):
        entry_list = [ent for ent in entry_list if os.path.exists(os.path.abspath(ent + data_suffix))]
        
    # ------------------------------------------------------------------------
    # if we've gotten here, we can proceed with normal combined data summary 

    # variables to put in summary page
    summary_heading_xlsx = summary_heading[:-2]
    # variables to put in events page
    events_heading_xlsx = events_heading

    # ---------------------------------
    # INITIALIZE WORKBOOK
    # ---------------------------------
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_events = wb.create_sheet("Events")

    # add column headers to pages
    ws_summary.append(summary_heading_xlsx)
    ws_events.append(events_heading_xlsx)

    # -------------------------------------------------------
    # LOOP THROUGH DATA AND WRITE TO XLSX FILE
    # -------------------------------------------------------
    for ent in entry_list:
        filename_full = ent + data_suffix
        if not os.path.exists(os.path.abspath(filename_full)):
            # in case we don't have analysis for this file yet
            print(ent + ' not yet analyzed--failed to save')
        else:
            # ------------------------------------------------------------
            # load combined data file
            filepath, filename = os.path.split(filename_full)
            flyCombinedData = hdf5_to_flyCombinedData(filepath, filename)
            filename_noExt, _ = os.path.splitext(filename)

            # check that data was loaded successfully
            if not bool(flyCombinedData):
                print('*Error loading {} -- skipping'.format(filename_full))
                continue

            # get bank and channel info
            xp_regex = "XP\d\d"
            channel_regex = "channel_\d"
            xp_name = re.findall(xp_regex, filename_noExt)[0]
            channel_name = re.findall(channel_regex, filename_noExt)[0]

            # -------------------------------------------
            # load in relevant bout data (feeding)
            bouts = flyCombinedData['bouts']
            # bouts = np.asarray(bouts)
            volumes = flyCombinedData['volumes']
            channel_t = flyCombinedData['channel_t']

            try:
                num_meals = bouts.shape[1]
            except IndexError:
                num_meals = 0
            total_volume = np.sum([float(vol) for vol in volumes])

            if (num_meals < 1):
                latency = -1.0
                duration_eating = 0.0
                bout_start_t = np.nan
                bout_end_t = np.nan

            else:
                bout_start_t = [channel_t[bouts[0, ith]] for ith in range(num_meals)]
                bout_end_t = [channel_t[bouts[1, ith]] for ith in range(num_meals)]
                latency = bout_start_t[0]
                duration_eating = np.sum(np.asarray(bout_end_t) - np.asarray(bout_start_t))

            # ---------------------------------------------------------
            # load in relevant kinematics data (tracking)

            cum_dist = flyCombinedData['cum_dist']
            vel_mag = flyCombinedData['vel_mag']
            moving_ind = flyCombinedData['moving_ind']
            vid_t = flyCombinedData['t']
            dist_mag = np.sqrt(flyCombinedData['xcm_smooth'] ** 2 +
                               flyCombinedData['ycm_smooth'] ** 2)

            avg_speed = np.nanmean(vel_mag[moving_ind])
            cum_dist_max = cum_dist[-1]
            frac_moving = float(np.sum(moving_ind)) / float(moving_ind.size)

            # ---------------------------------------------------------------
            # also get leave times, fraction of time spent in food zone 
            # (before and after first meal), and distance walked prior to first
            # meal
            dwell_times, censoring = get_dwell_time(bouts, channel_t,
                                                    dist_mag, vid_t,
                                                    fz_rad=fz_rad)

            if (num_meals < 1):
                pre_meal_pathlength = cum_dist_max
                fz_idx = (dist_mag <= fz_rad)
                fz_frac_pre = float(np.sum(fz_idx)) / float(dist_mag.size)
                fz_frac_post = 0.0
            else:
                first_meal_ind = bouts[0, 0]
                pre_meal_pathlength = cum_dist[first_meal_ind - 1]
                fz_idx = (dist_mag <= fz_rad)
                pre_ind = (vid_t < bout_start_t[0])
                post_ind = (vid_t >= bout_start_t[0])
                fz_frac_pre = float(np.sum(fz_idx & pre_ind)) / float(dist_mag.size)
                fz_frac_post = float(np.sum(fz_idx & post_ind)) / float(dist_mag.size)
            # ---------------------------------------------------------
            # WRITE DATA TO XLSX FILE
            # ---------------------------------------------------------
            # summary info (one line per fly)
            row_list = [filename_noExt, xp_name, channel_name,
                        num_meals, total_volume, duration_eating,
                        latency, cum_dist_max, avg_speed, frac_moving,
                        pre_meal_pathlength, fz_frac_pre, fz_frac_post]

            ws_summary.append(row_list)

            # event info (multiple lines per fly)
            for jth in range(num_meals):
                bout_start_curr = bout_start_t[jth]
                bout_end_curr = bout_end_t[jth]
                volume_curr = float(volumes[jth])
                duration_curr = bout_end_curr - bout_start_curr

                row_curr = [filename_noExt, xp_name, channel_name, jth + 1,
                            bout_start_curr, bout_end_curr, duration_curr,
                            volume_curr, float(dwell_times[jth]),
                            float(censoring[jth])]
                ws_events.append(row_curr)

                # -------------------------------------
    # SAVE XLSX FILE
    # -------------------------------------           
    wb.save(xlsx_filename)
    print('Completed saving {}'.format(xlsx_filename))


# ------------------------------------------------------------------------------
# function to save summary of combined data to HDF5 format
def save_comb_summary_hdf5(entry_list, h5_filename,
                           fz_rad=analysisParams['food_zone_rad']):
    print('Saving to {} ...'.format(h5_filename))
    # info for converting file types
    data_suffix = '_COMBINED_DATA.hdf5'

    # ------------------------------------------------------------------------
    # first check if the combined data exists -- if not, save whatever summary
    # we can (i.e. just tracking data)
    justTrackFlag = False
    comb_data_idx = [os.path.exists(os.path.abspath(ent + data_suffix)) for
                     ent in entry_list]
    if not any(comb_data_idx):
        # check for tracking data
        track_suffix = '_TRACKING_PROCESSED.hdf5'
        track_data_idx = [os.path.exists(os.path.abspath(ent + track_suffix))
                          for ent in entry_list]

        # if we find tracking data, save just tracking summary
        if any(track_data_idx):
            print('No feeding data -- saving just tracking data')
            data_suffix = track_suffix
            justTrackFlag = True
        else:
            # if neither combined nor tracking data, we can't save summary
            print('No tracking or combined data -- exiting')
            return

    # ---------------------------------
    # initialize HDF5 file
    # ---------------------------------
    counter = 1
    with h5py.File(h5_filename, 'w') as f:

        # -------------------------------------------------------
        # LOOP THROUGH DATA AND WRITE TO HDF5 FILE
        # -------------------------------------------------------
        for ent in entry_list:
            # current filename and the group name we'll assign it for hdf5 file
            filename_full = ent + data_suffix
            grp_name = "Fly{0:03d}".format(counter)
            # ------------------------------------------------
            # check if current file exists -- if not, skip
            if not os.path.exists(os.path.abspath(filename_full)):
                # in case we don't have analysis for this file yet
                print(ent + ' not yet analyzed--failed to save')
                continue

            # ------------------------------------------------------------
            # load combined data file
            filepath, filename = os.path.split(filename_full)
            if justTrackFlag:
                flyCombinedData = hdf5_to_flyTrackData(filepath, filename)
            else:
                flyCombinedData = hdf5_to_flyCombinedData(filepath, filename)
            filename_noExt, _ = os.path.splitext(filename)

            # check that data was loaded successfully
            if not bool(flyCombinedData):
                print('*Error loading {} -- skipping'.format(filename_full))
                continue

            # get bank and channel info
            xp_regex = "XP\d\d"
            channel_regex = "channel_\d"
            xp_name = re.findall(xp_regex, filename_noExt)[0]
            channel_name = re.findall(channel_regex, filename_noExt)[0]

            # -------------------------------------------
            # load in relevant bout data (feeding)
            if not justTrackFlag:
                bouts = flyCombinedData['bouts']
                # bouts = np.asarray(bouts)
                volumes = flyCombinedData['volumes']
                channel_t = flyCombinedData['channel_t']

                try:
                    num_meals = bouts.shape[1]
                except IndexError:
                    num_meals = 0
                total_volume = np.sum([float(vol) for vol in volumes])

                if (num_meals < 1):
                    latency = -1.0
                    duration_eating = 0.0
                    bout_start_t = np.nan
                    bout_end_t = np.nan
                    bout_dur = np.nan
                else:
                    bout_start_t = [channel_t[bouts[0, ith]] for ith in
                                    range(bouts.shape[1])]
                    bout_end_t = [channel_t[bouts[1, ith]] for ith in
                                  range(bouts.shape[1])]
                    bout_dur = [(te - ts) for (ts, te) in
                                zip(bout_start_t, bout_end_t)]
                    latency = bout_start_t[0]
                    duration_eating = np.sum(np.asarray(bout_end_t) -
                                             np.asarray(bout_start_t))

            # ---------------------------------------------------------
            # load in relevant kinematics data (tracking)

            cum_dist = flyCombinedData['cum_dist']
            vel_mag = flyCombinedData['vel_mag']
            moving_ind = flyCombinedData['moving_ind']
            vid_t = flyCombinedData['t']
            dist_mag = np.sqrt(flyCombinedData['xcm_smooth'] ** 2 +
                               flyCombinedData['ycm_smooth'] ** 2)

            avg_speed = np.nanmean(vel_mag[moving_ind])
            cum_dist_max = cum_dist[-1]
            frac_moving = float(np.sum(moving_ind)) / float(moving_ind.size)

            # ---------------------------------------------------------------
            # also get leave times, fraction of time spent in food zone 
            # (before and after first meal), and distance walked prior to
            # first meal
            if not justTrackFlag:
                dwell_times, censoring = get_dwell_time(bouts, channel_t, dist_mag, vid_t, fz_rad=fz_rad)

                if (num_meals < 1):
                    pre_meal_pathlength = cum_dist_max
                    fz_idx = (dist_mag <= fz_rad)
                    fz_frac_pre = float(np.sum(fz_idx)) / float(dist_mag.size)
                    fz_frac_post = 0.0
                else:
                    first_meal_ind = bouts[0, 0]
                    pre_meal_pathlength = cum_dist[first_meal_ind - 1]
                    fz_idx = (dist_mag <= fz_rad)
                    pre_ind = (vid_t < bout_start_t[0])
                    post_ind = (vid_t >= bout_start_t[0])
                    fz_frac_pre = float(np.sum(fz_idx & pre_ind)) / float(dist_mag.size)
                    fz_frac_post = float(np.sum(fz_idx & post_ind)) / float(dist_mag.size)

            # ---------------------------------------------------------
            # WRITE DATA TO HDF5 FILE
            # ---------------------------------------------------------
            # summary info
            f.create_dataset('{}/f_name'.format(grp_name), data=filename_noExt)
            f.create_dataset('{}/xp_name'.format(grp_name), data=xp_name)
            f.create_dataset('{}/ch_name'.format(grp_name), data=channel_name)

            # just tracking info
            f.create_dataset('{}/cum_dist_max'.format(grp_name),
                             data=cum_dist_max)
            f.create_dataset('{}/avg_vel_mag'.format(grp_name), data=avg_speed)
            f.create_dataset('{}/frac_moving'.format(grp_name),
                             data=frac_moving)

            # tracking info (time series)
            f.create_dataset('{}/t'.format(grp_name), data=vid_t)
            f.create_dataset('{}/cum_dist_ts'.format(grp_name), data=cum_dist)
            f.create_dataset('{}/dist_mag_ts'.format(grp_name), data=dist_mag)

            if not justTrackFlag:
                # feeding info
                f.create_dataset('{}/num_meals'.format(grp_name),
                                 data=num_meals)
                f.create_dataset('{}/tot_vol'.format(grp_name),
                                 data=total_volume)
                f.create_dataset('{}/dur_eating'.format(grp_name),
                                 data=duration_eating)
                f.create_dataset('{}/eat_latency'.format(grp_name),
                                 data=latency)

                # feeding info (mealwise)
                f.create_dataset('{}/meal_num'.format(grp_name),
                                 data=np.arange(num_meals))
                f.create_dataset('{}/start_time'.format(grp_name),
                                 data=np.asarray(bout_start_t))
                f.create_dataset('{}/end_time'.format(grp_name),
                                 data=np.asarray(bout_end_t))
                f.create_dataset('{}/meal_dur'.format(grp_name),
                                 data=np.asarray(bout_dur))
                f.create_dataset('{}/meal_vol'.format(grp_name),
                                 data=np.asarray(volumes))

                # tracking + feeding info 
                f.create_dataset('{}/pre_meal_dist'.format(grp_name),
                                 data=pre_meal_pathlength)
                f.create_dataset('{}/fz_frac_pre'.format(grp_name),
                                 data=fz_frac_pre)
                f.create_dataset('{}/fz_frac_post'.format(grp_name),
                                 data=fz_frac_post)
                f.create_dataset('{}/dwell_time'.format(grp_name),
                                 data=np.asarray(dwell_times))
                f.create_dataset('{}/dwell_censor'.format(grp_name),
                                 data=np.asarray(censoring, dtype=bool))

            # increment counter for next group
            counter += 1

    # notify user of completion
    print('Completed saving {}'.format(h5_filename))


# ----------------------------------------------------------------------------------------------------
# function to save turn rate in user-defined period surrounding a meal
# NB: should maybe try to generalize this, but not sure how many similar instances there will be
def save_meal_aligned_turn_rate(basic_entries, save_filename, window_left_sec=0.0, window_right_sec=10.0, meal_num=0,
                                turn_vel_thresh=analysisParams['turn_vel_thresh'],
                                turn_ang_thresh=analysisParams['turn_ang_thresh']):

    # suffix for filename with both feeding and tracking data
    data_suffix = '_COMBINED_DATA.hdf5'

    # we'll need to read x and y data to calculate turn rate
    varx = 'xcm_smooth'
    vary = 'ycm_smooth'

    # ---------------------------------------------------------------
    # initialize save output
    # ---------------------------------------------------------------
    data_all = []
    has_data_flag = []

    # ----------------------------------------------------
    # loop over data files
    for ith, ent in enumerate(basic_entries):
        # get full filename for current hdf5 analysis file, as well as file id
        hdf5_filename = ent + data_suffix
        if os.path.exists(hdf5_filename):
            ent_id = os.path.basename(ent)
            # read out meal aligned x, y, and t data (t is overkill, but need to get sampling freq)
            data_x = get_meal_aligned_data(hdf5_filename, varx, window_left_sec=window_left_sec,
                                           window_right_sec=window_right_sec, meal_num=meal_num)
            data_y = get_meal_aligned_data(hdf5_filename, vary, window_left_sec=window_left_sec,
                                           window_right_sec=window_right_sec, meal_num=meal_num)
            data_t = get_meal_aligned_data(hdf5_filename, 't', window_left_sec=window_left_sec,
                                           window_right_sec=window_right_sec, meal_num=meal_num)
        else:
            data_x = None
            data_y = None
            data_t = None

        # make sure we got data for both x and y
        if data_x is None or data_y is None:
            print('Failed to load data for {} -- skipping'.format(hdf5_filename))
            has_data_flag.append(False)
            continue

        # use x and y info to calculate turn rate in this window
        # (first detect turns, then take that number divided by duration of window)
        dt = np.mean(np.diff(data_t))
        turn_pts = detect_turn_points(data_x, data_y, dt=dt, thresh=turn_vel_thresh, min_turn_angle=turn_ang_thresh)

        turn_rate = np.size(turn_pts)/np.abs(window_right_sec - window_left_sec)

        # add current turn rate data to "all" list
        data_all.append(turn_rate)
        has_data_flag.append(True)  # also keep track of the fact that this file worked

    # --------------------------------------------------------------------
    # write combined data to workbook
    # --------------------------------------------------------------------

    # initialize workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Meal-aligned turn rate'

    # write meal number in first row
    meal_num_heading = ['Meal Number:', str(meal_num + 1)]  # NB: adding one to account for Python indexing
    ws.append(meal_num_heading)

    # write column headers in second row
    data_heading = ['Filename', 'Turn rate (turns/s)']
    ws.append(data_heading)

    # remove files in which there is no data (valid entries are ones with data)
    # basic_entries_valid = [ent for (kth, ent) in enumerate(basic_entries) if has_data_flag[kth]]
    basic_entries_valid = [ent for (tf, ent) in zip(has_data_flag, basic_entries) if tf]

    # from here, loop over files and write the turn rate as a row
    for (ent, trn_rate) in zip(basic_entries_valid, data_all):
        # construct row ( [filename, turn rate] ) and append to worksheet
        ws.append([os.path.basename(ent), trn_rate])

    # set column width to be wider (more readable)
    for col in ws.columns:
        col_str = get_column_letter(col[0].column)
        ws.column_dimensions[col_str].width = 22
    # save workbook
    wb.save(save_filename)

    return


# -------------------------------------------------------------------------------------
# NB: the following 3 functions have to do with the calculation of turn rate (above)
# -------------------------------------------------------------------------------------
# function to calculate 2d trajectory curvature
def calc_curvature(x, y, dt=1.0):
    # get derivatives
    x_dot = (1.0 / dt) * np.insert(np.diff(x), 0, 0.0)
    y_dot = (1.0 / dt) * np.insert(np.diff(y), 0, 0.0)

    x_ddot = (1.0 / dt) * np.insert(np.diff(x_dot), 0, 0.0)
    y_ddot = (1.0 / dt) * np.insert(np.diff(y_dot), 0, 0.0)

    # suppress warnings for divide by zero, since we'll almost certainly get them
    invalid_err_setting = np.geterr()
    np.seterr(invalid='ignore')

    # calculate speed
    vel_mag = np.sqrt(x_dot ** 2 + y_dot ** 2)

    # calculate curvature (formula from wikipedia)
    kappa = np.abs(x_dot * y_ddot - y_dot * x_ddot) / (vel_mag ** 3)

    # return warning to normal
    np.seterr(invalid=invalid_err_setting['invalid'])

    # return curvature
    return kappa


# -------------------------------------------------------------------------------------
# function to get angle of vector between successive points in 2d trajectory
def get_traj_angles(x, y):
    # take arctan of vectors between successive points (traja also restricts to positive vals)
    # ang = np.abs(np.arctan(np.diff(y)/np.diff(x)))
    ang = np.unwrap(np.arctan2(np.diff(y), np.diff(x)))

    # to keep array lengths the same, double up the first value
    return np.insert(ang, 0, ang[0])


# -------------------------------------------------------------------------------------
# function to detect turns in 2d trajectory (based on angular velocity and curvature)
def detect_turn_points(x, y, dt=1.0, thresh=1.2, min_turn_angle=None):
    # first calculate step by step angle and curvature (curvature used later)
    ang = get_traj_angles(x, y)
    # ang_full = np.insert(np.arctan(np.diff(y)/np.diff(x)),0,0.0)  # also get angle vals that include negative range
    kappa = calc_curvature(x, y)

    # get derivative of this wrt to time
    theta_diff = (1.0 / dt) * np.diff(ang)  # diff in angle -- SIGNED
    # theta_diff = (1.0 / dt) * np.abs(np.diff(ang))  # diff in angle

    # insert zero to keep arrays the same length
    theta_diff = np.insert(theta_diff, 0, 0)

    # find where theta_diff exceeds threshold value
    # NB: testing out separating by cw and ccw turns to keep turning periods from being merged
    turn_idx_cw = np.where((theta_diff > thresh), np.full(x.shape, True), np.full(x.shape, False))
    turn_idx_ccw = np.where((-1 * theta_diff > thresh), np.full(x.shape, True), np.full(x.shape, False))

    # get list whose entries are indices of putative turns
    turn_idx_list_cw = idx_by_thresh(turn_idx_cw)
    turn_idx_list_ccw = idx_by_thresh(turn_idx_ccw)

    # handle cases when we don't have any turns, only turns of one direction, etc
    if (turn_idx_list_cw is None) and (turn_idx_list_ccw is None):
        print('Could not detect any turns')
        return np.array([])
    elif (turn_idx_list_cw is None) and ~(turn_idx_list_ccw is None):
        turn_idx_list = turn_idx_list_ccw
    elif ~(turn_idx_list_cw is None) and (turn_idx_list_ccw is None):
        turn_idx_list = turn_idx_list_cw
    else:
        turn_idx_list = turn_idx_list_cw + turn_idx_list_ccw

    turn_idx_list.sort(key=lambda s: s[0])   # sort list entries by first index

    # check for false positives by looking at the difference between initial and final angles
    # turn_idx_list = idx_by_thresh(turn_idx)
    turn_angles = []
    for tidx in turn_idx_list:
        start_ind = np.max(np.array([tidx[0] - 1, 0]))
        end_ind = np.min(np.array([tidx[-1] + 1, ang.size - 1]))
        turn_angles.append(np.abs(ang[end_ind] - ang[start_ind]))

    if min_turn_angle:
        turn_idx_list = [turn_idx_list[ith] for ith in range(len(turn_idx_list)) if (turn_angles[ith] > min_turn_angle)]

    # now reduce list of indices in vicinity of turn to just one index (the turn point)
    # NB: this is determined by peak in curvature (could i just have skipped right to this?)
    turn_pts = [idx[np.argmax(kappa[idx])] for idx in turn_idx_list]

    # return these detected turn points
    return np.asarray(turn_pts)  # , turn_idx_list, turn_angles

